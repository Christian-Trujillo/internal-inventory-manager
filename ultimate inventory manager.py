from cmath import nan
from matplotlib.pyplot import show
import openpyxl as xl
import os, re
from googleapiclient.discovery import build
from google.oauth2 import service_account
import PySimpleGUI as sg
import pandas as pd
import numpy as np
import json
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from calendar import month, weekday
from datetime import datetime

### define variables needed for other funcs ###
def Initialize():
    global SCOPES
    global SERVICE_ACCOUNT_FILE
    global creds
    global service
    global CONTAINERS_ID
    global TRANSFERS_ID
    global INV_SAFETY_ID
    global FORECAST_ID
    global path
    global sheet
    global sku_list
    global sku_dict
    global packages
    global dates
    global period_to_weeks
    global daterange
    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    path = os.getcwd()
    # get info from keys.json
    SERVICE_ACCOUNT_FILE = path+r'\keys.JSON'
    creds= None
    creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes = SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    # The ID and range of a sample spreadsheet.
    CONTAINERS_ID = '10xCGUnnH9M8dLOhfCDak-ymbr1Vviz8CBJkRcsosHss'
    TRANSFERS_ID = '180OiU7000XFu3iN6BL0Edj1iKwmTEgfJ26xbT3nzoqk'
    INV_SAFETY_ID = '1yUyKPL4K7Pu2IysYXoMHoPbg3Y25jfQDr4shk1Dc3CA'
    FORECAST_ID = '1ehzKr6KoeLnzGjagddU9pZ-Feu7QITBWQO7LzrqOoCw'
    # Call the Sheets API
    sheet = service.spreadsheets()
    
    ### forecast variables ###
    
    with open('references.json','r+') as f:
        data=json.load(f) 
        sku_dict = data['sku_dict']
        packages = data['packages']
        sku_list = data['sku_list']
    ### list of dates to use in sales data ###
    def daterange(start_date, end_date):
        for n in range(int((end_date - start_date).days)):
            yield start_date + timedelta(n)
    
    period_to_weeks={'1 Week':1,'2 Weeks':2,'1 Month':4,'3 Months':13,'6 Months':23,'9 Months':39}

### pull live data from needed google sheets, return as a pandas DataFrame ###
def read_sheets():
    containers = pd.DataFrame(sheet.values().get(spreadsheetId=CONTAINERS_ID,range="'Current Containers'!A3:z100").execute().get('values',[]))
    inv_safety = pd.DataFrame(sheet.values().get(spreadsheetId=INV_SAFETY_ID,range="'inventory and safeties'!A2:d200").execute().get('values',[]))
    transfers = pd.DataFrame(sheet.values().get(spreadsheetId=TRANSFERS_ID,range="'Warehouse Transfers'!A2:M200",valueRenderOption = 'FORMULA').execute().get('values',[]))
    containers = containers.replace([None],['']).values.tolist()
    transfers = transfers.replace([None],['']).values.tolist()
    inv_safety = inv_safety.replace([None],['']).values.tolist()
    containers.remove(containers[0])
    transfers.remove(transfers[0])
    return containers,inv_safety, transfers
### enter str in format of sku1=qty1 / sku2=qty2/... . Returns dictionary of each container and their item qtys ###
def grab_qty(str):
    ### regex list for finding items in containers ###
    regexlist = []
    for item in sku_list:
        regexlist.append(rf'{item}=\d+')
        regexlist.append(rf'{item}= \d+')
        regexlist.append(rf'{item} =\d+')
        regexlist.append(rf'{item} = \d+')
    mydict={}
    xlist =re.findall(r"(?=("+'|'.join(regexlist)+r"))", str)
    for i in range(len(xlist)):
        try: xlist[i]=xlist[i].split('=')
        except: xlist.remove(xlist[i])
    for item in xlist:
        mydict[item[0]]=item[1]
    return mydict
### enter sku as item, uses grab_qty func to return dict of containers that contain the specific sku, {container#1:qty1, container#2, qty2...}
### only returns containers that have not been recieved and added into SC ###
def search_containers(item):
    containers_with_item = {}
    for row in containers:
        if row[15]=='':
            item_qty = grab_qty(row[11])
            if item.upper() in item_qty:
                containers_with_item[row[0]]=item_qty[item.upper()]
    return containers_with_item
### inputs data from containers recieved at storage warehouses into transfers page in speific format ###
def update_transfer():
    containers_updated=[]
    B=0
    M=0
    for i in range(len(transfers)):
        if transfers[i][0] == '':
            benson_len = i 
            break
    for i in range(len(transfers)):
        if transfers[i][7] == '':
            mag_len = i
            break
    for i in range(len(containers)):
        item_qty = grab_qty(containers[i][11])
        if containers[i][14].upper() == 'X' and containers[i][12].upper().find('UPDATED') == -1 and  containers[i][12].find('BROOKS') == -1:
            for key in item_qty:
                if containers[i][12].upper().find('BENSON') != -1:
                    transfers[benson_len+B][0] = str(containers[i][9])[:8]
                    transfers[benson_len+B][1] = key
                    transfers[benson_len+B][2] = '-'+item_qty[key]
                    transfers[benson_len+B][3] = containers[i][0]
                    containers[i][12] = 'BENSON - UPDATED'
                    containers[i][17]= ''
                    containers_updated.append(containers[i][0])
                    B+=1
                if containers[i][12].upper().find('MAGNOLIA') != -1:
                    transfers[mag_len+M][7] = str(containers[i][9])[:8]
                    transfers[mag_len+M][8] = key
                    transfers[mag_len+M][9] = '-'+item_qty[key]
                    transfers[mag_len+M][10] = containers[i][0]
                    containers[i][12] = 'MAGNOLIA - UPDATED'
                    containers[i][17]= ''
                    containers_updated.append(containers[i][0])
                    M+=1
    sheet.values().update(spreadsheetId=TRANSFERS_ID, range="'Warehouse Transfers'!A3", valueInputOption='USER_ENTERED', body={'values':transfers}).execute()
    sheet.values().update(spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A4", valueInputOption='USER_ENTERED', body={'values':containers}).execute()
    return containers_updated
### uses regex to find most recent exports for LOW INV REPORT and Safety export, if none are manually selected ###
def find_general_inv_files(inv = '', sfty=''):
    if inv == '' or sfty =='':
        path =  os.getcwd()
        sfty_list = []
        low_inv_list = []
        #regex file names
        regex_sfty = re.compile(r'Orders_Export_\d{6}.xlsx$')
        regex_low_inv = re.compile(r'LOW INVENTORY REPORT \(\d+\).xlsx$')
        for root,dirs,files in os.walk(path):
            for file in files:
                if regex_sfty.match(file):
                    sfty_list.append(file)
                if regex_low_inv.match(file):
                    low_inv_list.append(file)
        # set paths for worksheets by using highest numbered report
        low_inv= f'{path}\{max(low_inv_list)}'
        sfty = f'{path}\{max(sfty_list)}'
        # create workbook objects
        low_inv_wkbk = xl.load_workbook(low_inv)
        sfty_wkbk = xl.load_workbook(sfty)
        #create workbook active sheets object
        low_inv_obj=pd.DataFrame(low_inv_wkbk.active.values)
        sfty_obj=pd.DataFrame(sfty_wkbk.active.values)

    if inv != '':
        low_inv_wkbk = xl.load_workbook(inv)
        low_inv_obj=pd.DataFrame(low_inv_wkbk.active.values)
    if sfty != '':
        sfty_wkbk = xl.load_workbook(sfty)
        sfty_obj=pd.DataFrame(sfty_wkbk.active.values)

    return low_inv_obj,sfty_obj
### inputs data from inventory files from find_general_inv_files into general inventory google sheets ###
def update_inv_safety(low_inv_obj, sfty_obj):
    ### create dictionary for safety/min qty per sku ###
    sheet = service.spreadsheets()
    safety_list = sheet.values().get(spreadsheetId=INV_SAFETY_ID,range="'inventory and safeties'!h2:j66").execute().get('values',[])
    safety_dict={}
    for item in safety_list:
        safety_dict[item[0]]=[int(item[1]),int(item[2])]
    
    ### current inventories and safeties as df###
    low_inv = low_inv_obj
    sfty=sfty_obj
    ### 
    inv = pd.DataFrame(sheet.values().get(spreadsheetId=INV_SAFETY_ID,range="'inventory and safeties'!a2:f111").execute().get('values',[]))
    inv = inv.filter([0,5])
    index = pd.DataFrame(sku_list)
    sfty = sfty.filter([62,63])
    sfty.columns=[0,1]
    sfty=sfty[1:]
    sfty[1]=sfty[1].astype(int)
    low_inv = low_inv[1:]
    low_inv = low_inv.astype({2:int, 3:int})

    result = pd.merge(index, low_inv.filter([0,2,3]), on=0)
    result = pd.merge(result, sfty, how="left", on=0)
    result[5]=''
    result = pd.merge(result, inv, how="left", on=0)
    result.columns=['SKU', 'AGGREGATE','PHYSICAL','SAFETY', 'ERRORS','NOTES']

    for sku in sku_list:
        if sku not in safety_dict.keys():
            safety_dict[sku]=[0,0]
    safety_dict=pd.DataFrame(safety_dict,index=['safety','min']).transpose()
    result.index=result['SKU']
    safety_dict=safety_dict.loc[result.index.tolist(),:]
    ### fill errors col ###   
    conditions = [result["PHYSICAL"]<result["SAFETY"],result["AGGREGATE"]<0,(result["PHYSICAL"]>result['SAFETY']) & (result['SAFETY']<safety_dict['safety']),result['SAFETY']<safety_dict['safety'],result['SAFETY']<safety_dict['min']]
    choices = ['Physical Inventory Incorrect;\n please Cycle Count','Negative Aggregate;\n check backorders and Cycle Count','Item may be back in stock;\n please Cycle Count','Safety Reduced','Minimum Reduced']
    # result['ERRORS'] = np.where(result["PHYSICAL"]<result["SAFETY"], 'Physical Inventory Incorrect;\n please Cycle Count',result['ERRORS'])
    # result['ERRORS'] = np.where(result["AGGREGATE"]<0, 'Negative Aggregate;\n check backorders and Cycle Count',result['ERRORS'])
    # result['ERRORS'] = np.where((result["PHYSICAL"]>result['SAFETY'] )& (result['SAFETY']<=safety_dict[result['SKU']][0]), 'Item may be back in stock;\n please Cycle Count',result['ERRORS'])
    # result['ERRORS'] = np.where(result['SAFETY']<safety_dict[result['SKU']][0], 'Safety Reduced',result['ERRORS'])
    # result['ERRORS'] = np.where(result['SAFETY']<safety_dict[result['SKU']][1], 'Minimum Reduced',result['ERRORS'])
    result['ERRORS'] = np.select(conditions, choices, default='')

    result = result.replace([None],[''])
    result = result.replace([nan],['']) 
    result['SAFETY']=result['SAFETY'].replace([''],[0])
    result=[result.columns.tolist()] + result.values.tolist()


    sheet.values().update(spreadsheetId=INV_SAFETY_ID, range=f'inventory and safeties!a1', valueInputOption='USER_ENTERED', body={'values':result}).execute()
### editing and adding of containers into containers sheets ###
def update_containers(container_number, freight_forwarder,ETA, contents, MNFCR, Notes_1,Notes_2):
    
    container_numbers = {}
    for i in range(len(containers)):
        container_numbers[containers[i][0]]=i
    if container_number not in container_numbers:
        containers.append([container_number, freight_forwarder, '','',ETA,'','','','','','',contents,'',MNFCR,'','','',Notes_1,Notes_2,'','','','','',''])
        sheet.values().update(spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A4", valueInputOption='USER_ENTERED', body={'values':containers}).execute()

    elif container_number in container_numbers:
        popup = sg.Window('Overwrite Container Log?',[[sg.Text('Container is already in log')],[sg.Text('Overwrite With Entered Values?')],[sg.Button('yes'),sg.Button('no')]])
        event,values = popup.read()
        if event =='yes':
            row = containers[container_numbers[container_number]]
            new_row = [container_number, freight_forwarder, '','',ETA,'','','','','','',contents,'',MNFCR,'','','',Notes_1,Notes_2,'','','','','','']
            for i in range(len(new_row)):
                if new_row[i]=='' and row[i]!='':
                    new_row[i]=row[i]
            containers[container_numbers[container_number]] = new_row
            popup.close()
            sheet.values().update(spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A4", valueInputOption='USER_ENTERED', body={'values':containers}).execute()

        if event =='no':
            popup.close()
### returns DataFrame of totals for low_inv_report AGG + on water inv (from search_containers) ###
def item_quantity():

    item_qty={}
    for row in inv_safety:
        item_qty[row[0]]=int(row[1])
    for item in sku_list:
        total=0
        if search_containers(item)!={}:
            dict = search_containers(item)
            for qty in dict.values():
                total+=int(qty)
            item_qty[item]+=total
    df= pd.DataFrame(item_qty, index=['Stock']).transpose()
    return df
### totals sales of items in sku_list, per day, oer a period of time (dates list). Returns in DataFrame ###
def sales(path,method):
    ### create lsit of dates to go over based on forecasting method ###
    today = datetime.today()
    method = [method]
    if method ==['both']:
        method = ['Avg Delta','Seasonality']
    if 'Avg Delta' in method:
        start_date = today- relativedelta(years=1, months = 7)
    else:
        start_date = today- relativedelta(months = 7)
    while weekday(start_date.year,start_date.month,start_date.day)!=0:
        start_date+=relativedelta(days=1)
    end_date = today
    dates =[]
    for x in daterange(start_date, end_date):
        dates.append(x.strftime("%m/%d/%Y"))

    SC = pd.read_excel(path).filter(['Ship Date', 'SKU', 'Qty Sold'])
    ### reformat date ###}'
    window['process status'].update(value='processing...\naccumulating sales...') 
    window.refresh()
    SC.sort_values('Ship Date')
    ship_dict = {}
    nonlist=[]
    for sku in sku_list:
        ship_dict[sku] = {}
        for x in dates:
            ship_dict[sku][x] = 0
    for row in SC.itertuples(index=False):
        if row[0] not in dates:
            continue
        if row[1] in sku_dict.keys():
            for i in sku_dict[row[1]]:
                ship_dict[i][row[0]] += row[2]
        elif row[1] in sku_list:
            ship_dict[row[1]][row[0]] += row[2]
        else: 
            if row[1] not in nonlist:
                nonlist.append(row[1])
    SC = pd.DataFrame(ship_dict)
    window['process status'].update(value='processing...\nsales accumulated!')
    window.refresh()        
    #return SC.reindex(index=SC.index[::-1]).transpose()
    return SC.transpose()
### searches for sections of 10 days without sales in an item, replaces given day with aerage over last 30 days ###
def remove_OOS(df):
    window['process status'].update(value='processing...\nsales accumulated!\nremoving OOS periods...')
    window.refresh()
    df_modified = df.copy()
    for sku in range(len(df.index)):
        sku_avg=[]
        for i in range(30,df.shape[1]):
            if df.loc[df.index[sku],f'{df.columns[i]}']>0:
                sku_avg.append(df.loc[df.index[sku],f'{df.columns[i]}'])
        for i in range(30,df.shape[1]):
            if df.loc[df.index[sku],f'{df.columns[i]}']==0:
                avg_qty=[]
                check_foward_dates=[]
                check_back_dates=[]
                for n in range(1,10):
                    try:check_foward_dates.append(df.loc[df.index[sku],f'{df.columns[i+n]}'])
                    except:pass
                    try:check_back_dates.append(df.loc[df.index[sku],f'{df.columns[i-n]}'])
                    except:pass
                if (len(check_foward_dates)==9 and sum(check_foward_dates)==0 ) or (sum(check_back_dates)==0 and len(check_back_dates)==9):
                    for n in range(1,31):
                        avg_qty.append(df_modified.iloc[sku,i-n])
                    if np.average(avg_qty)==0:
                        df_modified.loc[df.index[sku],f'{df.columns[i]}'] = np.average(sku_avg)
                    else: df_modified.loc[df.index[sku],f'{df.columns[i]}']= np.average(avg_qty)
    window['process status'].update(value='processing...\nsales accumulated!\nOOS periods removed!')
    window.refresh()        

        #### use for fixing OOS->0 sales days  ###
    # for sku in range(len(Inv.index)):
    #     for i in range(Inv.shape[1]-180,Inv.shape[1]):
    #         if Inv.loc[Inv.index[sku],f'{Inv.columns[i]}']==0:
    #             avg_qty=[]
    #             for n in range(1,31):
    #                 avg_qty.append(S.iloc[sku,i-n])
    #             S_modified.loc[Inv.index[sku],f'{Inv.columns[i]}']= np.average(avg_qty)  
    return df_modified
### collects dates in sales DF into 7-day periods starting on sundays, returns DF of each weeks total sales, named by first date in week ###
def weekify(df):
    window['process status'].update(value='processing...\nsales accumulated!\nOOS periods removed!\ncollecting sales into weeks...')
    window.refresh()
    df1=pd.DataFrame()
    m=0
    n=1
    weeks=[]
    weeks_dict={}
    cols1 = {}
    columns =df.columns.tolist()
    for i in range(1,114):
        exec(f'week_{i} = []')

    for column in columns:
        if m==0:
            cols1[n]=column
        exec(f'week_{n}.append(column)')
        m+=1
        
        if m==7:
            exec(f'weeks.append(week_{n})')
            exec(f'weeks_dict["week_{n}"]=week_{n}')
            n+=1
            m=0
    
    n=0
    for week in weeks:
        n+=1
        exec(f'df1["{cols1[n]}"] = 0')
        exec(f'df1["{cols1[n]}"]=df[{week}].sum(axis=1)')
    window['process status'].update(value='processing...\nsales accumulated!\nOOS periods removed!\nsales collected into weeks!')   
    return df1.transpose()#, weeks_dict
### returns either a running average over a given period for each week ###
### OR returns the change in sales from period to period (delta). Special rules for longer periods)
### OR returns both ###
def running_avg(df,period = 1, avg = True, delta=True):# periods of 1,2,3,4,13,26,39 weeks
    lst=[]
    lst2 = []
    space = period
    length = 3*period
    if period>25: 
        space = 13
        length = period+26
    Avg=pd.DataFrame()
    Delta=pd.DataFrame()
    if avg:
        for sku in range(len(df.columns)):
            for i in range(0,df.shape[0]-(length)):
                lst=[]
                if space<12:
                    for j in range(0,length, space):
                        lst2=[]
                        for k in range(period):
                            lst2.append(df.iloc[i+j+k,sku])
                        lst.append(np.sum(lst2))
                    Avg.loc[df.index[i+length],f'{df.columns[sku]}'] = np.round(np.average(lst),0)
                elif space>12:
                    for j in range(0,39, space):
                        lst2=[]
                        for k in range(period):
                            lst2.append(df.iloc[i+j+k,sku])
                        lst.append(np.sum(lst2))
                    Avg.loc[df.index[i+length],f'{df.columns[sku]}'] = np.round(np.average(lst),0)
    if delta:
        for sku in range(len(df.columns)):
            for i in range(0,df.shape[0]-1):
                Delta.loc[df.index[i+1],f'{df.columns[sku]}']= df.iloc[i+1,sku]- df.iloc[i,sku]
    # delta gives chage of sales
    # delta2 gives change in slope (negative means not rising as fast)
    if avg and delta:
        return Avg, Delta
    elif avg:
        return Avg
    elif delta:
        return Delta
### returns average change in sales over last 3 periods, added to last periods sales ###
def DF_forecast(self,period):
    delta= running_avg(running_avg(self, period, avg=False), period, delta=False).iloc[-1]
    sales= self.iloc[-1:-(period+1):-1].sum()
    return sales.add(delta,fill_value=0).transpose()
### returns sales DF , replacing each week with its total sales over the next period length ###
def DF_grouping(self, period):
    lst=[]
    Avg=pd.DataFrame()
    
    for sku in range(len(self.columns)):
        for i in range(0,self.shape[0]-period):
            lst=[]
            for j in range(0,period):
                lst.append(self.iloc[i+j,sku])
            Avg.loc[self.index[i],f'{self.columns[sku]}'] = np.sum(lst)
    return Avg   
### given 1 year 6 months of sales data, returns each quartes sales as a percentage of the previous quarters sales ###
def seasonality_model(df,period=13):
    
    Percentage=pd.DataFrame()
    
    for sku in range(len(df.columns)):
        for i in range(0,df.shape[0]-2*period):
            lst=[]
            lst2=[]
            for j in range(0,period):
                lst2.append(df.iloc[i+j,sku])
            for j in range(period,2*period):
                lst.append(df.iloc[i+j,sku])
            Percentage.loc[df.index[i+period],f'{df.columns[sku]}'] = np.sum(lst)/np.sum(lst2)
    return Percentage 
### uses seasonality_model for projecting 3 months of sales to forecast longer periods, assuming similiar seasonality trends and 2021 ###
def project_Seasonality(self,S_percentage, period):
    s_n=self.grouping(13)
    projection=pd.DataFrame()
    ### for every data point ###
    for sku in range(len(self.columns)):
        for i in range(self.shape[0]-1,self.shape[0]):
            ### returns index in S_pertentage closest to Self's index ###
            time=self.index[i]
            index=S_percentage.index.searchsorted(time)
            try: time2=S_percentage.index[index]
            except: time2=S_percentage.index[0]
            time3=S_percentage.index[index-1]
            
            
            if np.abs(date(2021,int(time2[:2]),int(time2[3:5]))-date(2021,int(time[:2]),int(time[3:5])))<np.abs(date(2021,int(time3[:2]),int(time3[3:5]))-date(2021,int(time[:2]),int(time[3:5]))):
                a=index
                b=a+13
                c=a+26
            else: 
                a=index-1
                b=a+13
                c=a+26
            ### checks that index values for S_percentage are within S_percentages's index 
            if a>=len(S_percentage.index):
                a=a-len(S_percentage.index)
            if b>=len(S_percentage.index):
                b=b-len(S_percentage.index)
            if c>=len(S_percentage.index):
                c=c-len(S_percentage.index)
            ### multplies last 3 months sales by seasonality muliplier to return next 3 months sales ###
            projection.loc[self.index[i],f'{self.columns[sku]}'] = S_percentage.loc[S_percentage.index[a],f'{S_percentage.columns[sku]}']*s_n.loc[s_n.index[i-13],f'{s_n.columns[sku]}']
            ### adds next 3 month period -> 6 months ###
            if period==26:
                projection.loc[self.index[i],f'{self.columns[sku]}'] += projection.loc[self.index[i],f'{self.columns[sku]}'] *S_percentage.loc[S_percentage.index[b],f'{S_percentage.columns[sku]}']
            ### adds next 6 month period -> 9 months ###
            if period==39:
                projection.loc[self.index[i],f'{self.columns[sku]}'] += (projection.loc[self.index[i],f'{self.columns[sku]}'] *S_percentage.loc[S_percentage.index[b],f'{S_percentage.columns[sku]}'])+(projection.loc[self.index[i],f'{self.columns[sku]}'] *S_percentage.loc[S_percentage.index[b],f'{S_percentage.columns[sku]}']*S_percentage.loc[S_percentage.index[c],f'{S_percentage.columns[sku]}'])
    return projection
pd.DataFrame.project=project_Seasonality
pd.DataFrame.grouping=DF_grouping
pd.DataFrame.cast=DF_forecast
### formats different length forecasts into single DF ###
### adds columns for predicted order qtys from factory based on 6 months sales ###
### writes data to forecasting google sheet, along with sales data over given periods of time ###
def Process_Forecast(Sales,method, for_export=True):
    method=[method]
    if method==['both']:
        method=['Avg Delta','Seasonality']
    ### format and print onto excel sheet
    ### 2 forecast for seperate methods of  forecasting longer periods (3m,6m,9m) ###
    forecast_df1=pd.DataFrame()
    forecast_df2=pd.DataFrame()
    forecast_list1=[]
    forecast_list2=[]
    S = Sales.copy()
    ### set up to calculate dales extra sales due to time between now and order date ###
    S_model= pd.read_json(path+r'\seasonality.json')
    day = date(int(S.index[-1][6:]), int(S.index[-1][:2]),int(S.index[-1][3:5]))+timedelta(weeks=1)
    next_month=day.month+1
    orderdate = date(day.year, next_month, 15)
    orderdate2 = orderdate+relativedelta(months=1)
    weeks_difference=(orderdate-day).days/7
    ### totals sales, fill OOS dates, totals to weeks, then takes year off of dates in index ###
    
    new_index=[]
    for i in range(len(S.index)):
        new_index.append(S.index[i][:5])
    S.index=new_index
    
    inventory=item_quantity()
    week_cast=S.cast(1).transpose().apply(lambda x: round(x,0))
    two_week_cast=S.cast(2).transpose().apply(lambda x: round(x,0))
    month_cast=S.cast(4).transpose().apply(lambda x: round(x,0))
    if 'Avg Delta' in method:
        three_month_cast=S.cast(13).transpose().apply(lambda x: round(x,0))
        six_month_cast=S.cast(26).transpose().apply(lambda x: round(x,0))
        nine_month_cast=S.cast(39).transpose().apply(lambda x: round(x,0))
        filler_qty1 = (nine_month_cast-six_month_cast)*(weeks_difference/13)
        forecast_df1['1 Week']=week_cast
        forecast_df1['2 Weeks']=two_week_cast
        forecast_df1['1 Month']=month_cast
        forecast_df1['3 Months']=three_month_cast
        forecast_df1['6 Months']=six_month_cast
        forecast_df1['9 Months']=nine_month_cast
        ### assign forecast to named columns in df ###
        for period in forecast_df1.columns:
            forecast_df1[period][forecast_df1[period]<0]=0
        forecast_df1['']=''
        forecast_df1['Stock']=inventory
        ### deficits 6 months out from order date ###
        forecast_df1[f'{orderdate.strftime("%m/%d")} Order QTY']=(forecast_df1['6 Months']+filler_qty1-forecast_df1['Stock']).apply(lambda x: round(x,0))
        forecast_df1[f'{orderdate2.strftime("%m/%d")} Order QTY']=((forecast_df1['9 Months']-forecast_df1['6 Months'])/3).apply(lambda x: round(x,0))
        forecast_df1=forecast_df1.fillna(0)
        ### convert df to list with column and index values ###
        forecast_list1 = [forecast_df1.columns.tolist()] + forecast_df1.reset_index().values.tolist()
        forecast_list1[0].insert(0,str(S.index[-1:][0]))

    if 'Seasonality' in method:
        three_month_project=S.project(S_model,13).transpose().apply(lambda x: round(x,0))
        six_month_project=S.project(S_model,26).transpose().apply(lambda x: round(x,0))
        nine_month_project=S.project(S_model,39).transpose().apply(lambda x: round(x,0))
        filler_qty2 = (nine_month_project-six_month_project)*(weeks_difference/13)
        forecast_df2['1 Week']=week_cast
        forecast_df2['2 Weeks']=two_week_cast
        forecast_df2['1 Month']=month_cast
        forecast_df2['3 Months']=three_month_project
        forecast_df2['6 Months']=six_month_project
        forecast_df2['9 Months']=nine_month_project
        ### assign forecast to named columns in df ###
        for period in forecast_df2.columns:
            forecast_df2[period][forecast_df2[period]<0]=0
        forecast_df2['']=''
        forecast_df2['Stock']=inventory
        ### deficits 6 months out from order date ###
        forecast_df2[f'{orderdate.strftime("%m/%d")} Order QTY']=(forecast_df2['6 Months']+filler_qty2.transpose()-forecast_df2['Stock']).transpose().apply(lambda x: round(x,0))
        forecast_df2[f'{orderdate2.strftime("%m/%d")} Order QTY']=((forecast_df2['9 Months']-forecast_df2['6 Months'])/3).apply(lambda x: round(x,0))
        forecast_df2=forecast_df2.fillna(0)
        ### convert df to list with column and index values ###
        forecast_list2 = [forecast_df2.columns.tolist()] + forecast_df2.reset_index().values.tolist()
        forecast_list2[0].insert(0,str(S.index[-1:][0]))
    if for_export:
        ### create different forecasts ###
        s2=S.grouping(2).transpose().apply(lambda x: round(x,0))
        s4=S.grouping(4).transpose().apply(lambda x: round(x,0))
        s13=S.grouping(13).transpose().apply(lambda x: round(x,0))

        group1=[S[-5:].transpose()[-5:].columns.tolist()] + S[-5:].transpose().reset_index().values.tolist()
        group2 = [s2.iloc[:,-5:].columns.tolist()] + s2.iloc[:,-5:].reset_index().values.tolist()
        group4 = [s4.iloc[:,-5:].columns.tolist()] + s4.iloc[:,-5:].reset_index().values.tolist()
        group13 = [s13.iloc[:,-5:].columns.tolist()] + s13.iloc[:,-5:].reset_index().values.tolist()
            
        group1[0].insert(0,'')
        group2[0].insert(0,'')
        group4[0].insert(0,'')
        group13[0].insert(0,'')
        if 'Avg Delta' in method:
            s26=S.grouping(26).transpose().apply(lambda x: round(x,0))
            s39=S.grouping(39).transpose().apply(lambda x: round(x,0))
            group26 = [s26.columns.tolist()] + s26.reset_index().values.tolist()
            group39 = [s39.columns.tolist()] + s39.reset_index().values.tolist()
            group26[0].insert(0,'')
            group39[0].insert(0,'')
            return [forecast_list1,forecast_list2, group1, group2, group4, group13,group26,group39]
        else:  return [forecast_list1,forecast_list2, group1, group2, group4, group13,[],[]]
    else: 
        return [forecast_df1,forecast_df2]
def Export_Forecast(exports , method):
    method=[method]
    if method==['both']:
        method=['Avg Delta','Seasonality']
    forecast_list1=exports[0]
    forecast_list2=exports[1]
    group1=exports[2]
    group2=exports[3]
    group4=exports[4]
    group13=exports[5]
    group26=exports[6]
    group39=exports[7]
    if 'Avg Delta' in method:
        sheet.values().update(spreadsheetId=FORECAST_ID, range=f'Forecast method 1!a1', valueInputOption='USER_ENTERED', body={'values':forecast_list1}).execute()
    if 'Seasonality' in method:
        sheet.values().update(spreadsheetId=FORECAST_ID, range=f'Forecast method 2!a1', valueInputOption='USER_ENTERED', body={'values':forecast_list2}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'1 W history!a1', valueInputOption='USER_ENTERED', body={'values':group1}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'2 W history!a1', valueInputOption='USER_ENTERED', body={'values':group2}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'1 M history!a1', valueInputOption='USER_ENTERED', body={'values':group4}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'3 M history!a1', valueInputOption='USER_ENTERED', body={'values':group13}).execute()
    if 'Avg Delta' in method:
        sheet.values().update(spreadsheetId=FORECAST_ID, range=f'6 M history!a1', valueInputOption='USER_ENTERED', body={'values':group26}).execute()
        sheet.values().update(spreadsheetId=FORECAST_ID, range=f'9 M history!a1', valueInputOption='USER_ENTERED', body={'values':group39}).execute()
def Forecast_graphing(sales,forecast,period,sku):
    P = period_to_weeks[period]
    N=forecast.loc[sku,period]
    n1=sales[sku][-1:]
    n1_date=date(2022,int(sales.index[-1][:2]),int(sales.index[-1][3:5]))
    dates_list=[]
    values_list=[]
    for T in range(P+1):
        dates_list.append((n1_date+relativedelta(weeks=T)).strftime("%m/%d/%Y"))
        values_list.append(n1+((2*T/P)*((N/P)-n1)))
    return pd.concat([sales[sku],pd.DataFrame(values_list,index=dates_list)],join='outer')
def process_sales():
    try:
        window['process status'].update(value='Processing...')
        window.refresh()
        processed_sales = weekify(remove_OOS(sales(values['Sales Export File'],values['method'])))
        window['process status'].update(value='sales accumulated!\nOOS periods removed!\nsales collected into weeks!\n-- Processed! --')   
        window.Refresh()
        return processed_sales
    except: window['process status'].update(value='Please choose a Sales Export and a forecasting method')
### set GUI layout and size ###
Initialize()
sg.theme('DarkAmber')
main_layout=[[sg.Text('Main Menu')],[sg.Button('General Inventory', size= 190)],[sg.T('')],[sg.Button('Warehouse Transfers', size= 190)],[sg.T('')],[sg.Button('Containers Logs', size= 190)],[sg.T('')],[sg.Button('Sales Forecast', size= 190)],[sg.T('')]]
### exec(layout_txt) to recreate laout variables becasue they are not reuasable ###
general_inv_layout=[]
warehouse_transfers_layout=[]
containers_logs_layout=[]
forecast_layout=[]
containers_search_layout=[]
graph_forecast_layout=[]
main_layout_txt="main_layout=[[sg.Text('Main Menu')],[sg.Button('General Inventory', size= 190)],[sg.T('')],[sg.Button('Warehouse Transfers', size= 190)],[sg.T('')],[sg.Button('Containers Logs', size= 190)],[sg.T('')],[sg.Button('Sales Forecast', size= 190)],[sg.T('')]]"
general_inv_layout_txt = "general_inv_layout=[[sg.Button('back'),sg.Text('General Inventory')],[sg.Text('Inventory Sheet: '),sg.Text(key = 'inv_file_in'),sg.FileBrowse('Select Inventory Sheet', key = 'inv_file_out')],[sg.Text('Safety Qty Export: '),sg.Text(key='sfty_file_in'),sg.FileBrowse('Select Safety Qty Export', key = 'sfty_file_out')],[sg.T('')],[sg.Button('update', size = 190)],[sg.Text(key='finished')]]"
warehouse_transfers_layout_txt="warehouse_transfers_layout=[[sg.Button('back'),sg.Text('Warehouse Transfers')],[sg.Button('Update Containers to Warehouse Transfers', size = 190)],[sg.Text('Containers: '),sg.Text(key='containers update list')]]"
containers_logs_layout_txt="containers_logs_layout = [[sg.Button('back'),sg.Text('Containers Logs')],[sg.T('')],[sg.Text('Container Number'),sg.Text('                                                  '),sg.Text('Freight Forwarder'),sg.Text('                                                  '),sg.Text('ETA')],[sg.Input(key='Container Number'),sg.Input(key='Freight Forwarder'),sg.Input(key='ETA')],[sg.Text('Contents')],[sg.Input(key='Contents')],[sg.Text('MNFCR'),sg.Text('                                                                '),sg.Text('Management Notes'),sg.Text('                                               '),sg.Text('Additional Notes')],[sg.Input(key='MNFCR'),sg.Input(key='Management Notes'),sg.Input(key='Additional Notes')],[sg.T('')],[sg.Button('Add Container'), sg.Button('Search Containers')]]"
forecast_layout_txt ="forecast_layout=[[sg.Button('back'),sg.Text('Sales Forecast')],[sg.Text('Sales Export:'),sg.Text(key = 'Sales Export'),sg.FileBrowse('Browse',key='Sales Export File')],[sg.Combo(['both', 'Avg Delta','Seasonality'],key='method'),sg.Button('graph & compare',key='graph'),sg.Button('Export',key='Export'),sg.Text(key='process status')],[sg.T('')],[]]"
containers_search_layout_txt = "containers_search_layout =[[sg.Button('back'),sg.Text('Search Containers')],[sg.Text('search sku: '), sg.Input(key='sku'),sg.Button('Search')],[sg.Text(key = 'containers'),sg.Text('             '),sg.Text(key='QTY')]]"
graph_forecast_layout_txt="graph_forecast_layout = [[sg.Text('SKU:'),sg.Combo(sku_list,key='SKU')],[sg.T('Period:'),sg.Combo(['1 Week','2 Weeks','1 Month','3 Months','6 Months','9 Months'],key='Period'),sg.B('Graph')]]"
window=sg.Window('Internal Inventory Manager', main_layout, size=(800,250))
### how window interacts with given inputs ###
while True:
    containers,inv_safety, transfers = read_sheets()
    event,values = window.read()
    print(event)

    if event == sg.WIN_CLOSED:
        break

### document the rest of the interface ###
    elif event == 'General Inventory':  ## General inventory page ##
        ### redefine layout, close main, open general. read events , enable back button ###
        Submenu = True
        exec(general_inv_layout_txt)
        window.close()
        window= sg.Window('Internal Inventory Manager', general_inv_layout, size=(800,200))
        while Submenu == True:
            event,values = window.read()
            containers,inv_safety, transfers = read_sheets()
            print(event)
            if event == sg.WIN_CLOSED:
                break
            if event == 'back':
                Submenu=False
            ## general specific events ##
            if values['inv_file_out'] !='' or values['sfty_file_out']!='':
                name_1 = values['inv_file_out']
                window['inv_file_in'].update(value=name_1)
                name_2 = values['sfty_file_out']
                window['sfty_file_in'].update(value=name_2)
            if event == 'update':
                try:
                    low_inv_obj,sfty_obj= find_general_inv_files(inv=values['inv_file_out'],sfty= values['sfty_file_out'])
                    update_inv_safety(low_inv_obj,sfty_obj) 
                    window['finished'].update(value='Done')
                except:
                    
                    window['finished'].update(value='-- ERROR --')

        window.close()
        exec(main_layout_txt)        
        window=sg.Window('Internal Inventory Manager', main_layout, size=(800,250))
        continue
    
    elif event == 'Warehouse Transfers': ## warehouse transfers page ##
        ### redefine layout, close main, open general. read events , enable back button ###
        Submenu = True
        exec(warehouse_transfers_layout_txt)
        window.close()
        window= sg.Window('Internal Inventory Manager', warehouse_transfers_layout, size=(800,200))
        while Submenu == True:
            event,values = window.read()
            containers,inv_safety, transfers = read_sheets()
            print(event)
            if event == sg.WIN_CLOSED:
                break
            if event == 'back':
                Submenu=False
            if event == 'Update Containers to Warehouse Transfers':
                container_str = ''
                container_list = update_transfer() 
                for code in container_list:
                    container_str+=code+',  '
                
                window['containers update list'].update(value=container_str)
        window.close()
        exec(main_layout_txt)
        window=sg.Window('Internal Inventory Manager', main_layout, size=(800,250))
        continue

    elif event == 'Containers Logs':  ## containers logs page ##
            ### redefine layout, close main, open general. read events , enable back button ###
            Submenu = True
            exec(containers_logs_layout_txt)
            window.close()
            window= sg.Window('Internal Inventory Manager', containers_logs_layout, size=(800,300))
            container_numbers = []
            for row in containers:
                container_numbers.append(row[0])
            while Submenu == True:
                event,values = window.read()
                containers,inv_safety, transfers = read_sheets()
                print(event)
                if event == sg.WIN_CLOSED:
                    break
                if event == 'back':
                    Submenu=False
                if event =='Add Container':
                    update_containers(values['Container Number'].upper(),values['Freight Forwarder'].upper(),values['ETA'].upper(),values['Contents'].upper(),values['MNFCR'].upper(),values['Management Notes'],values['Additional Notes'])
                if event =='Search Containers':
                    Subsubmenu = True
                    exec(containers_search_layout_txt)
                    window.close()
                    window= sg.Window('Internal Inventory Manager', containers_search_layout, size=(800,300))
                    SubMenu=False
                    while Subsubmenu == True:
                        event,values = window.read()
                        containers,inv_safety, transfers = read_sheets()
                        if event == sg.WIN_CLOSED:
                            break
                        if event == 'back':
                            Submenu=True
                            Subsubmenu=False
                            window.close()
                            exec(containers_logs_layout_txt)
                            window=sg.Window('Internal Inventory Manager', containers_logs_layout, size=(800,300))
                            continue
                        if event == 'Search':
                            index = search_containers(values['sku'].upper())
                            found_containers = ''
                            found_QTYs = ''
                            for container in index:
                                found_containers+=container+'\n\n'
                                found_QTYs+=index[container]+'\n\n'
                            window['containers'].update(value=found_containers)
                            window['QTY'].update(value=found_QTYs)

            window.close()
            exec(main_layout_txt)
            window=sg.Window('Internal Inventory Manager', main_layout, size=(800,250))
            continue

    elif event == 'Sales Forecast':   ## sales forecasting page ##
        ### redefine layout, close main, open general. read events , enable back button ###
        Submenu = True
        exec(forecast_layout_txt)
        window.close()
        window= sg.Window('Internal Inventory Manager', forecast_layout, size=(800,250))
        while Submenu == True:
            event,values = window.read()
            containers,inv_safety, transfers = read_sheets()
            print(event)
            if event == sg.WIN_CLOSED:
                break
            if event == 'back':
                Submenu=False
            ## general specific events ##
            if event == "Export" or event == 'graph':
                if (values['method'] not in ['Avg Delta','Seasonality','both']):
                    continue
                if event == 'Export':
                    processed_sales=process_sales()
                    Export_Forecast(Process_Forecast(processed_sales,values['method'],for_export=True),values['method'])
                    # except:window['process status'].update(value='Please choose a Sales Export:\n   -1yr 7m for Avg Delta\n   -4 months for Seasonality')
                if event == 'graph':
                    processed_sales=process_sales()
                    forecasts = Process_Forecast(processed_sales,values['method'],False)
                    Subsubmenu = True
                    exec(graph_forecast_layout_txt)
                    window.close()
                    window= sg.Window('Internal Inventory Manager', graph_forecast_layout, size=(800,300))
                    SubMenu=False
                    while Subsubmenu == True:
                        event,values = window.read()
                        containers,inv_safety, transfers = read_sheets()
                        if event == sg.WIN_CLOSED:
                            break

                        if event =='Graph':
                            ### for given method, sku and period, plot sales next to idealized forecast ###
                            ax=processed_sales[values['SKU']].plot()
                            for forecast in forecasts:
                                if forecast.empty == False:
                                    Forecast_graphing(processed_sales,forecast, values['Period'],values['SKU']).plot(ax=processed_sales[values['SKU']].plot())
                                
                            show()

        window.close()
        exec(main_layout_txt)        
        window=sg.Window('Internal Inventory Manager', main_layout, size=(800,250))
        continue
window.close()

  ### grab all containers with specific item ###  
# find all containers containing item                                 X 
# list container numbers and quantities                               X
#  search with fedex api???



