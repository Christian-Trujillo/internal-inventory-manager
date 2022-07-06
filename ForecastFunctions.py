from cmath import nan
from matplotlib.pyplot import show
import openpyxl as xl
import os
import re
from googleapiclient.discovery import build
from google.oauth2 import service_account
import PySimpleGUI as sg
import pandas as pd
import numpy as np
import json
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from calendar import month, weekday
from datetime import datetime
from time import sleep
from pyparsing import alphanums, alphas
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import shutil
import re
from webdriver_manager.chrome import ChromeDriverManager
from threading import Thread
import pickle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
###* define variables needed for other funcs ###


def Initialize(queue):
    '''Sets several GLOBAL variables , as well as importing information from several internal logs\n\n
        returns variables to main file if used through a module'''

    global SCOPES
    global SERVICE_ACCOUNT_FILE
    global creds
    global service
    global CONTAINERS_ID
    global TRANSFERS_ID
    global INV_SAFETY_ID
    global FORECAST_ID
    global path
    global sheet
    global sku_list
    global sku_dict
    global reduced_sku_list
    global sku_details
    global period_to_weeks
    global daterange
    global forecast_list
    global today
    global num_period
    global period_num
    global credentials

    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    path = os.getcwd()
    # get info from keys.json
    SERVICE_ACCOUNT_FILE = path+r'\keys.JSON'
    creds = None
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    # client id = '318827037112-eq655i5b2ns04g8pobdlbrch6g3nqu9u.apps.googleusercontent.com'
    service = build('sheets', 'v4', credentials=creds)
    # The ID and range of a sample spreadsheet.
    CONTAINERS_ID = '1d8hRkptQwV9VPLhgJLsVgYdwk8exz-GdF5oBdwjFbQE'
    TRANSFERS_ID = '1Di-2F9A4xSndZGOx7plq2JDzzConxMy0KRjeHO0a_BI'
    # TRANSFERS_ID = '180OiU7000XFu3iN6BL0Edj1iKwmTEgfJ26xbT3nzoqk'
    INV_SAFETY_ID = '1Li0W6bfRA-x80TyOt7yGus5y12AzMucom4NdrfWfqRU'
    FORECAST_ID = '13sJ9Iyp3rfeas7du6xwYeNLquORTE1OF75hQMKK7kdY'
    # Call the Sheets API
    sheet = service.spreadsheets()

    ### forecast variables ###

    with open('references.json', 'r+') as f:
        data = json.load(f)
        credentials = data['Login Credentials']
        sku_dict = data['sku_dict']
        reduced_sku_list = data['reduced_sku_list']
        sku_list = data['sku_list']
        sku_details = data['sku details']
    ### list of dates to use in sales data ###

    def daterange(start_date, end_date):
        for n in range(int((end_date - start_date).days)):
            yield start_date + timedelta(n)
    today = datetime.today().strftime("%y-%m-%d")
    period_to_weeks = {'1 Week': 1, '2 Weeks': 2, '1 Month': 4, '2 Months': 8, '3 Months': 13,
                       '4 Months': 17, '5 Months': 22, '6 Months': 26, '7 Months': 31, '8 Months': 35, '9 Months': 39}
    forecast_list = None
    num_period = {1: '1 Month', 2: '2 Months', 3: '3 Months', 4: '4 Months',
                  5: '5 Months', 6: '6 Months', 7: '7 Months', 8: '8 Months', 9: '9 Months',0:''}
    period_num = {'1 Month': 1, '2 Months': 2, '3 Months': 3, '4 Months': 4,
                  '5 Months': 5, '6 Months': 6, '7 Months': 7, '8 Months': 8, '9 Months': 9, '':0}

    queue.put((reduced_sku_list, sku_list, period_to_weeks,
              sku_details, num_period, period_num, credentials))
###* pull live data from needed google sheets, return as a pandas DataFrame ###


def read_sheets(queue):
    '''pulls data from internal logs regarding different details on current inventory status\n\n
        Returns Pandas DataFrames through queue'''
    global containers
    global transfers
    global inv_safety
    containers = pd.DataFrame(sheet.values().get(
        spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A3:z100").execute().get('values', []))
    inv_safety = pd.DataFrame(sheet.values().get(
        spreadsheetId=INV_SAFETY_ID, range="'inventory and safeties'!A2:d150").execute().get('values', []))
    transfers = pd.DataFrame(sheet.values().get(spreadsheetId=TRANSFERS_ID,
                             range="'Warehouse Transfers'!A2:T200", valueRenderOption='FORMULA').execute().get('values', []))
    containers = containers.replace([None], ['']).values.tolist()
    transfers = transfers.replace([None], ['']).values.tolist()
    inv_safety = inv_safety.replace([None], ['']).values.tolist()
    containers.remove(containers[0])
    transfers.remove(transfers[0])
    queue.put((containers, transfers, inv_safety))
###* enter str in format of sku1=qty1 / sku2=qty2/... . Returns dictionary of each container and their item qtys ###


def grab_qty(str):
    '''Given a string, Returns a dictionary of Cosmo skus and respective quantities\n\n
        primarily used for Cosmo container logs where the typical format is given by:\n
        \t"sku1= quantity1" / "sku2 =quantity2" / "sku3=quantity3" / etc...'''
    ### regex list for finding items in containers ###
    regexlist = []
    for item in sku_list:
        regexlist.append(rf'{item}=\d+')
        regexlist.append(rf'{item}= \d+')
        regexlist.append(rf'{item} =\d+')
        regexlist.append(rf'{item} = \d+')
    mydict = {}
    xlist = re.findall(r"(?=("+'|'.join(regexlist)+r"))", str.upper())
    for i in range(len(xlist)):
        try:
            xlist[i] = xlist[i].split('=')
        except:
            xlist.remove(xlist[i])
    for item in xlist:
        mydict[item[0]] = item[1]
    return mydict
#* enter sku as item, uses grab_qty func to return dict of containers that contain the specific sku, {container#1:qty1, container#2, qty2...}
###* only returns containers that have not been recieved and added into SC ###


def search_containers(item):
    '''uses grab_qty() function to programatically search container logs for COSMO SKUs\n\n
        Returns dictionary with format :
        \tcontainer number : item quantity'''
    containers_with_item = {}
    for row in containers:
        if row[15] == '':
            item_qty = grab_qty(row[11])
            if item.upper() in item_qty:
                containers_with_item[row[0]] = item_qty[item.upper()]
    return containers_with_item
###* inputs data from containers recieved at storage warehouses into transfers page in speific format ###


def update_transfer():
    '''Searches Cosmo Container Logs for received conatiners, searches contents using grab_qty(), marks updated\n
        then pastes items and quantities to respective warehouses in Transfer Logs\n\n
        Returns List of updated containers'''
    containers_updated = []
    B = 0
    M = 0
    O = 0
    for i in range(len(transfers)):
        if transfers[i][0] == '':
            benson_len = i
            break
    for i in range(len(transfers)):
        if transfers[i][7] == '':
            mag_len = i
            break
    for i in range(len(transfers)):
        if transfers[i][14] == '':
            ont_len = i
            break
    for i in range(len(containers)):
        item_qty = grab_qty(containers[i][11])
        if containers[i][14].upper() == 'X' and containers[i][12].upper().find('UPDATED') == -1 and containers[i][12].find('BROOKS') == -1:
            for key in item_qty:
                if containers[i][12].upper().find('BENSON') != -1:
                    transfers[benson_len+B][0] = str(containers[i][9])[:8]
                    transfers[benson_len+B][1] = key
                    transfers[benson_len+B][2] = '-'+item_qty[key]
                    transfers[benson_len+B][3] = containers[i][0]
                    containers[i][12] = 'BENSON - UPDATED'
                    containers[i][17] = ''
                    containers_updated.append(containers[i][0])
                    B += 1
                if containers[i][12].upper().find('MAGNOLIA') != -1:
                    transfers[mag_len+M][7] = str(containers[i][9])[:8]
                    transfers[mag_len+M][8] = key
                    transfers[mag_len+M][9] = '-'+item_qty[key]
                    transfers[mag_len+M][10] = containers[i][0]
                    containers[i][12] = 'MAGNOLIA - UPDATED'
                    containers[i][17] = ''
                    containers_updated.append(containers[i][0])
                    M += 1
                if containers[i][12].upper().find('ONTARIO') != -1:
                    transfers[ont_len+O][14] = str(containers[i][9])[:8]
                    transfers[ont_len+O][15] = key
                    transfers[ont_len+O][16] = '-'+item_qty[key]
                    transfers[ont_len+O][17] = containers[i][0]
                    containers[i][12] = 'ONTARIO - UPDATED'
                    containers[i][17] = ''
                    containers_updated.append(containers[i][0])
                    O += 1
    sheet.values().update(spreadsheetId=TRANSFERS_ID, range="'Warehouse Transfers'!A3",
                          valueInputOption='USER_ENTERED', body={'values': transfers}).execute()
    sheet.values().update(spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A4",
                          valueInputOption='USER_ENTERED', body={'values': containers}).execute()
    return containers_updated
###* uses regex to find most recent exports for LOW INV REPORT and Safety export, if none are manually selected ###


def find_general_inv_files(inv='', sfty=''):
    '''no longer needed\n\n
        used to find SC exports within folder\n
        Returns DataFrame of Inventory and Safety exports'''
    if inv == '' or sfty == '':
        path = os.getcwd()
        sfty_list = []
        low_inv_list = []
        # regex file names
        regex_sfty = re.compile(r'Orders_Export_\d{6}.xlsx$')
        regex_low_inv = re.compile(r'LOW INVENTORY REPORT \(\d+\).xlsx$')
        for root, dirs, files in os.walk(path):
            for file in files:
                if regex_sfty.match(file):
                    sfty_list.append(file)
                if regex_low_inv.match(file):
                    low_inv_list.append(file)
        # set paths for worksheets by using highest numbered report
        low_inv = f'{path}\{max(low_inv_list)}'
        sfty = f'{path}\{max(sfty_list)}'
        # create workbook objects
        low_inv_wkbk = xl.load_workbook(low_inv)
        sfty_wkbk = xl.load_workbook(sfty)
        # create workbook active sheets object
        low_inv_obj = pd.DataFrame(low_inv_wkbk.active.values)
        sfty_obj = pd.DataFrame(sfty_wkbk.active.values)

    if inv != '':
        low_inv_wkbk = xl.load_workbook(inv)
        low_inv_obj = pd.DataFrame(low_inv_wkbk.active.values)
    if sfty != '':
        sfty_wkbk = xl.load_workbook(sfty)
        sfty_obj = pd.DataFrame(sfty_wkbk.active.values)

    return low_inv_obj, sfty_obj
###* inputs data from inventory files from find_general_inv_files into general inventory google sheets ###


def update_inv_safety(low_inv_obj, sfty_obj):
    '''Given 2 DataFrames containing inventory and safety data:\n
        uses Item_quantity() to fill on_water quantity,\n
        fills Error column based on relationships between AGG. , PHYS. and SAFETY quantities,\n
        saves notes and reassigns back to respective items.\n
        \tUpdates Daily Inventory Report directly'''
    ### create dictionary for safety/min qty per sku ###
    sheet = service.spreadsheets()
    safety_list = sheet.values().get(spreadsheetId=INV_SAFETY_ID,
                                     range="'inventory and safeties'!h2:j66").execute().get('values', [])
    safety_dict = {}
    for item in safety_list:
        safety_dict[item[0]] = [int(item[1]), int(item[2])]

    ### current inventories and safeties as df###
    low_inv = low_inv_obj
    sfty = sfty_obj
    ###
    inv = pd.DataFrame(sheet.values().get(spreadsheetId=INV_SAFETY_ID,
                       range="'inventory and safeties'!a2:f111").execute().get('values', []))
    past_inv = inv.filter([0, 2]).astype({2: int})
    inv = inv.filter([0, 5])
    index = pd.DataFrame(sku_list)
    sfty = sfty.filter([63, 64])
    sfty[63][sfty[63] == 'COS-640SLTX-E'] = 'COS-640STX-E'
    sfty.columns = [0, 1]
    sfty = sfty[1:]
    sfty[1] = sfty[1].astype(int)
    low_inv = low_inv[1:]
    low_inv = low_inv.astype({2: int, 3: int})

    result = pd.merge(index, low_inv.filter([0, 2, 3]), on=0)
    result = pd.merge(result, sfty, how="left", on=0)
    result[5] = ''
    result = pd.merge(result, inv, how="left", on=0)
    result = pd.merge(result, past_inv, how="left", on=0)
    result.columns = ['SKU', 'AGGREGATE', 'PHYSICAL',
                      'SAFETY', 'ERRORS', 'NOTES', 'past inv']

    for sku in sku_list:
        if sku not in safety_dict.keys():
            safety_dict[sku] = [0, 0]
    safety_dict = pd.DataFrame(
        safety_dict, index=['safety', 'min']).transpose()
    result.index = result['SKU']
    # past_inv.index=past_inv[0]
    # past_inv.sort_index()
    safety_dict = safety_dict.loc[result.index.tolist(), :]
    result = result.replace([nan], [0])
    ### fill errors col ###
    conditions = [(result["AGGREGATE"] < 1) & (result['SAFETY'] > safety_dict['min']),
                  (result["PHYSICAL"] == 0) | ((result["AGGREGATE"] < 1)
                                               & (result['SAFETY'] <= safety_dict['min'])),
                  (result["PHYSICAL"] > result['past inv']) & (
                      result['SAFETY'] < safety_dict['safety']),
                  result["AGGREGATE"] < 0,
                  result['SAFETY'] < safety_dict['min'],
                  result['SAFETY'] < safety_dict['safety']]

    choices = ['Lower Safety; \n Item is not selling',
               'Item OOS',
               'Item may be back in stock;\n please Cycle Count',
               'Negative Aggregate;\n check backorders and Cycle Count',
               'Minimum Reduced', 'Safety Reduced']

    result['ERRORS'] = np.select(conditions, choices, default='')
    result.drop(['past inv'], 1, inplace=True)
    result["NOTES"].replace([0, '0'], '', inplace=True)
    result = result.replace([None], [''])
    result = result.replace([nan], [''])
    result['SAFETY'] = result['SAFETY'].replace([''], [0])
    result = [result.columns.tolist()] + result.values.tolist()

    sheet.values().update(spreadsheetId=INV_SAFETY_ID, range=f'inventory and safeties!G1',
                          valueInputOption='USER_ENTERED', body={'values': [[datetime.today().strftime('%B %e')]]}).execute()
    sheet.values().update(spreadsheetId=INV_SAFETY_ID, range=f'inventory and safeties!a1',
                          valueInputOption='USER_ENTERED', body={'values': result}).execute()
###* back in stock sheet ###


def update_back_in_stock(low_inv_obj):
    '''updates BIS sheet as a redundancy based on relaionships between Agg. , Phys. and Safety amounts'''
    sheet = service.spreadsheets()
    df = pd.DataFrame(sheet.values().get(spreadsheetId=INV_SAFETY_ID,
                      range="'inventory and safeties'!A2:C200").execute().get('values', []))
    df.index = df[0]
    df = df.drop([0, 2], axis=1)
    df.astype(int)
    low_inv_obj = low_inv_obj.drop(0)
    low_inv_obj.index = low_inv_obj[0]
    low_inv_obj = low_inv_obj.drop([0, 1, 3], axis=1)
    df = df.merge(low_inv_obj, left_on=df.index,
                  right_on=low_inv_obj.index, how='left')
    df.columns = ['SKU', 'Yesterday', 'Today']
    df.index = df['SKU']
    df = df.drop('SKU', axis=1)
    df = df.astype({'Yesterday': int, 'Today': int})
    df['B.I.S'] = ""
    df['B.I.S'][df['Yesterday'] < 5] = np.where(
        df['Today'] > df['Yesterday'], 'X', "")
    if not df['Today'].equals(df['Yesterday']):
        df_list = [df.columns.tolist()] + df.reset_index().values.tolist()
        df_list[0].insert(0, 'SKU')
        sheet.values().update(spreadsheetId=INV_SAFETY_ID, range=f"'back in stock (temp)'!a1",
                              valueInputOption='USER_ENTERED', body={'values': df_list}).execute()

###* editing and adding of containers into containers sheets ###


def update_containers(container_number, freight_forwarder, ETA, contents, MNFCR, Notes_1, Notes_2):
    ''' adds a row at the bottom of container logs with arguments placed in respective columns'''

    container_numbers = {}
    for i in range(len(containers)):
        container_numbers[containers[i][0]] = i
    if container_number not in container_numbers:
        containers.append([container_number, freight_forwarder, '', '', ETA, '', '', '', '',
                          '', '', contents, '', MNFCR, '', '', '', Notes_1, Notes_2, '', '', '', '', '', ''])
        sheet.values().update(spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A4",
                              valueInputOption='USER_ENTERED', body={'values': containers}).execute()

    elif container_number in container_numbers:
        popup = sg.Window('Overwrite Container Log?', [[sg.Text('Container is already in log')], [
                          sg.Text('Overwrite With Entered Values?')], [sg.Button('yes'), sg.Button('no')]])
        event, values = popup.read()
        if event == 'yes':
            row = containers[container_numbers[container_number]]
            new_row = [container_number, freight_forwarder, '', '', ETA, '', '', '', '', '',
                       '', contents, '', MNFCR, '', '', '', Notes_1, Notes_2, '', '', '', '', '', '']
            for i in range(len(new_row)):
                if new_row[i] == '' and row[i] != '':
                    new_row[i] = row[i]
            containers[container_numbers[container_number]] = new_row
            popup.close()
            sheet.values().update(spreadsheetId=CONTAINERS_ID, range="'Current Containers'!A4",
                                  valueInputOption='USER_ENTERED', body={'values': containers}).execute()

        if event == 'no':
            popup.close()
###* returns DataFrame of totals for low_inv_report AGG + on water inv (from search_containers) ###


def item_quantity():
    '''uses Search_containers() to search container logs for each Cosmo SKU\n
        Returns DF indexed by sku'''

    item_qty = {}
    for row in inv_safety:
        item_qty[row[0]] = int(row[1])
    for item in sku_list:
        total = 0
        if search_containers(item) != {}:
            dict = search_containers(item)
            for qty in dict.values():
                total += int(qty)
            item_qty[item] += total
    df = pd.DataFrame(item_qty, index=['Stock']).transpose()
    return df
###* downloads sales export and updates directory ###


def SC_login(driver):
    '''login to SellerCloud'''
    email = driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$txtEmail")
    pwd = driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$txtPwd")
    email.clear()
    pwd.clear()
    email.send_keys(credentials['Username'])
    pwd.send_keys(credentials['Pass']+Keys.ENTER)
###* downoad SC exports needed for updating sheets ###


def download_sales(chrome_path):
    '''Logs in and navigates through SellerCloud to export 1 year 7 months of sales for Forecasting\n
        saves file to exports/Sales Data.xlsx'''
    method = ['Avg Delta', 'Seasonality']
    if 'Avg Delta' in method:
        start_date = (datetime.today()-relativedelta(years=1,months=7)).strftime("%m/%d/%Y")
    else:
        start_date = (datetime.today()-relativedelta(months=4)).strftime("%m/%d/%Y")
    end_date = (datetime.today()).strftime("%m/%d/%Y")
    ### opens Chrome ###
    options = webdriver.ChromeOptions()
    options.headless = True
    prefs = {"download.default_directory": os.getcwd()+r'\exports'}
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(executable_path=chrome_path, chrome_options=options)
    ### looks up value in Chrome ###
    driver.get("https://df.cwa.sellercloud.com/DashboardV2/Reports/ReportV2_ProductQtySoldByDay.aspx")
    SC_login(driver)
    ### waits until title contains text ###
    WebDriverWait(driver, 30).until(EC.title_contains('Qty Sold By Product Per Day Report'))  # This is a dummy element
    ### checks if "text" in </title> ###
    assert "Qty Sold By Product Per Day Report - SellerCloud" == driver.title
    driver.find_element( By.NAME, "ctl00$ContentPlaceHolder1$txtFromDate").send_keys(start_date)
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$txtToDate").send_keys(end_date)
    run_report = driver.find_element(By.ID, "ContentPlaceHolder1_btnRunReport")
    run_report.click()

    # webdriver.ActionChains(driver).click(run_report).perform()
    # WebDriverWait(driver,180).until(EC.element_to_be_selected ((By.NAME, 'ctl00$ContentPlaceHolder1$imgExcel')))
    export_excel = driver.find_element( By.NAME, 'ctl00$ContentPlaceHolder1$imgExcel')
    export_excel.click()
    waiting_for_dl = True
    while waiting_for_dl:
        try:
            shutil.move(path +r'\exports\ProductQuantitySoldByDay.xlsx' , path + rf'\exports\Sales Data.xlsx')
            break
        except:
            sleep(1)
    driver.close()


def download_safeties(chrome_path):
    '''Logs in and navigates through SellerCloud to export Cosmo Safety Quantity Order\n
        saves file to exports/Safeties.xlsx'''
    options = webdriver.ChromeOptions()
    options.headless = True
    prefs = {"download.default_directory": os.getcwd()+r'\exports'}
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome( executable_path=chrome_path, chrome_options=options)
    driver.get("https://df.cwa.sellercloud.com/Orders/Orders_details.aspx?ID=7808146")
    SC_login(driver)
    WebDriverWait(driver, 30).until(EC.title_contains('Order 7808146'))
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$ActionList").send_keys('Export Order')
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$ImageButton1").click()
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$ddlFileType").send_keys('Excel')
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$btnExportOrders").click()
    regex = re.search(r'JobID=\d+', driver.page_source)
    JobId = regex.group()[6:]
    today = datetime.today().strftime("%m/%d/%Y")
    driver.get(f'https://df.cwa.sellercloud.com/MyAccount/QueuedJobs.aspx?UserID=2691946&JobType=-1&SubmittedOnStartDate={today}&SubmittedOnEndDate={today}&Status=-1')
    waiting_for_dl = True
    while waiting_for_dl:
        try:
            driver.find_element( By.ID, 'ContentPlaceHolder1_QueuedJobsList_grdMain_ctl00_ctl04_btnViewOutput').click()
            break
        except:
            sleep(10)
            driver.refresh()
    while waiting_for_dl:
        try:
            shutil.move( path + rf'\exports\Orders_Export_{JobId}.xlsx', path + r'\exports\Safeties.xlsx')
            break
        except:
            sleep(1)
    driver.close()


def download_inv(chrome_path):
    '''Logs in and navigates through SellerCloud to export Cosmo Inventory\n
        saves file to exports/Inventory.xlsx'''
    mod_sku_list = ''
    for sku in sku_list[:-1]:
        mod_sku_list += sku+' , '
    mod_sku_list += sku_list[-1]
    options = webdriver.ChromeOptions()
    options.headless = True
    prefs = {"download.default_directory": os.getcwd()+r'\exports'}
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(
        executable_path=chrome_path, chrome_options=options)
    driver.get("https://df.cwa.sellercloud.com/Inventory/ManageInventory.aspx?CompanyIDList=&sku=&active=1&rowsperPage=50&inventoryFrom=-2147483648&inventoryTo=-2147483648&SavedSearchName=&SKUUseWildCards=False&OrderID=0&InventoryViewMode=0&InventoryQtyFilterMode=0&SortBy=bvc_Product.ID&SortByDirection=ASC&")
    SC_login(driver)
    sleep(2)
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$txtSKU").send_keys(mod_sku_list)
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$btnSearchNow").click()
    driver.find_element(By.ID, "ContentPlaceHolder1_chkSelectAll").click()
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$btnExportProducts").click()
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$ddlExportformat").send_keys('Excel')
    driver.find_element(
        By.XPATH, "//a[@href='/inventory/CustomExport.aspx']").click()
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$ddlTemplate").send_keys('AVC AGG/PHYS')
    driver.find_element(By.ID, "ContentPlaceHolder1_btnLoadTemplate").click()
    driver.find_element(
        By.ID, "ContentPlaceHolder1_ddlExportFileFormat").click()
    sleep(1)
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$btnExport").click()
    sleep(3)
    regex = re.search(r'JobID=\d+', driver.page_source)
    JobId = regex.group()[6:]
    driver.get(
        f'https://df.cwa.sellercloud.com/MyAccount/QueuedJobs.aspx?UserID=2691946&JobType=-1&SubmittedOnStartDate={today}&SubmittedOnEndDate={today}&Status=-1')
    waiting_for_dl = True
    while waiting_for_dl:
        try:
            driver.find_element(
                By.ID, 'ContentPlaceHolder1_QueuedJobsList_grdMain_ctl00_ctl04_btnViewOutput').click()
            break
        except:
            sleep(10)
            driver.refresh()
    while waiting_for_dl:
        try:
            shutil.move(
                path + rf'\exports\\{JobId}.xlsx', path + r'\exports\Inventory.xlsx')
            break
        except:
            sleep(1)
    driver.close()


def download_vel(chrome_path):
    '''Logs in and navigates through SellerCloud to export Cosmo Sku Velocities\n
        saves file to exports/Velocities.xlsx'''
    mod_sku_list = ''
    for sku in reduced_sku_list[:-1]:
        mod_sku_list += sku+' , '
    mod_sku_list += sku_list[-1]
    options = webdriver.ChromeOptions()
    options.headless = True
    prefs = {"download.default_directory": os.getcwd()+r'\exports'}
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(
        executable_path=chrome_path, chrome_options=options)
    driver.get("https://df.cwa.sellercloud.com/Inventory/PredictedPurchasing.aspx")
    SC_login(driver)
    sleep(2)
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$txtProductID").send_keys(mod_sku_list)
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$ddlDaysOfOrder").send_keys(30)
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$ddlDaysToOrder").send_keys(30)
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$btnSearch").click()
    driver.find_element(By.NAME, "ctl00$ContentPlaceHolder1$ddlAction").send_keys(
        'Export to Excel')
    driver.find_element(
        By.NAME, "ctl00$ContentPlaceHolder1$btnDoAction").click()
    while True:
        try:
            shutil.move(path + rf'\exports\PredictPurchasing.xlsx',
                        path + r'\exports\Velocities.xlsx')
            break
        except:
            sleep(1)
    driver.close()
### totals sales of items in sku_list, per day, oer a period of time (dates list). Returns in DataFrame ###


def sales(path):
    '''Given Sales export from SC:\n
        sums item sales per day\n
        Returns DF with each Items sales per day\n
        \tIndex: SKU \n
        \tColumns: date'''
    ### create list of dates to go over based on forecasting method ###
    today = datetime.today()
    start_date = today - relativedelta(years=1, months=7)
    while weekday(start_date.year, start_date.month, start_date.day) != 0:
        start_date += relativedelta(days=1)
    end_date = today
    dates = []
    for x in daterange(start_date, end_date):
        dates.append(x.strftime("%m/%d/%Y"))

    SC = pd.read_excel(path).filter(['Ship Date', 'SKU', 'Qty Sold'])
    # reformat date ###}'
    SC.sort_values('Ship Date')
    ship_dict = {}
    nonlist = []
    for sku in reduced_sku_list:
        ship_dict[sku] = {}
        for x in dates:
            ship_dict[sku][x] = 0
    for row in SC.itertuples(index=False):
        if row[0] not in dates:
            continue
        if row[1] in sku_dict.keys():
            for i in sku_dict[row[1]]:
                ship_dict[i][row[0]] += row[2]
        elif row[1] in reduced_sku_list:
            ship_dict[row[1]][row[0]] += row[2]
        else:
            if row[1] not in nonlist:
                nonlist.append(row[1])
    SC = pd.DataFrame(ship_dict)
    # return SC.reindex(index=SC.index[::-1]).transpose()
    return SC.transpose()
### searches for sections of 10 days without sales in an item, replaces given day with average over last 30 days ###


def remove_OOS(df):
    '''Given processed saled DF:
        fills any 10+ day period with average sales '''
    s1 = df.copy().transpose()
    s1.roll_back = s1.rolling(10).sum() == 0
    s1.roll_forward = s1.iloc[::-1].rolling(10).sum().iloc[::-1] == 0
    s1.rolled = (s1.roll_back+s1.roll_forward)

    for index in range(len(s1.index)):
        for column in s1.columns:
            if s1.rolled.loc[s1.index[index], column] == True:
                if index <= 30:
                    s1.loc[s1.index[index], column] = s1[column].mean()
                else:
                    s1.loc[s1.index[index], column] = s1.loc[:s1.index[index], column].rolling(30).mean()[
                        index]
    return s1.transpose()
### collects dates in sales DF into 7-day periods starting on sundays, returns DF of each weeks total sales, named by first date in week ###


def weekify(df):
    '''groups processed sales into weeks, beginning  each monday, removes last, non-full week'''
    df1 = pd.DataFrame()
    m = 0
    n = 1
    weeks = []
    weeks_dict = {}
    cols1 = {}
    columns = df.columns.tolist()
    for i in range(1, 114):
        exec(f'week_{i} = []')

    for column in columns:
        if m == 0:
            cols1[n] = column
        exec(f'week_{n}.append(column)')
        m += 1

        if m == 7:
            exec(f'weeks.append(week_{n})')
            exec(f'weeks_dict["week_{n}"]=week_{n}')
            n += 1
            m = 0

    n = 0
    for week in weeks:
        n += 1
        exec(f'df1["{cols1[n]}"] = 0')
        exec(f'df1["{cols1[n]}"]=df[{week}].sum(axis=1)')
    return df1.transpose()
### returns either a running average over a given period for each week ###
# OR returns the change in sales from period to period (delta). Special rules for longer periods)
### OR returns both ###


def running_avg(df, period=1, avg=True, delta=True):  # periods of 1,2,3,4,13,26,39 weeks
    '''provides running average over given period, as well as change in sales\n\t
        can be stacked with itself to provide change in average or average change'''
    lst = []
    lst2 = []
    space = period
    length = 3*period
    if period > 25:
        space = 13
        length = period+26
    Avg = pd.DataFrame()
    Delta = pd.DataFrame()
    if avg:
        for sku in range(len(df.columns)):
            for i in range(0, df.shape[0]-(length)):
                lst = []
                if space < 12:
                    for j in range(0, length, space):
                        lst2 = []
                        for k in range(period):
                            lst2.append(df.iloc[i+j+k, sku])
                        lst.append(np.sum(lst2))
                    Avg.loc[df.index[i+length],
                            f'{df.columns[sku]}'] = np.round(np.average(lst), 0)
                elif space > 12:
                    for j in range(0, 39, space):
                        lst2 = []
                        for k in range(period):
                            lst2.append(df.iloc[i+j+k, sku])
                        lst.append(np.sum(lst2))
                    Avg.loc[df.index[i+length],
                            f'{df.columns[sku]}'] = np.round(np.average(lst), 0)
    if delta:
        for sku in range(len(df.columns)):
            for i in range(0, df.shape[0]-1):
                Delta.loc[df.index[i+1],
                          f'{df.columns[sku]}'] = df.iloc[i+1, sku] - df.iloc[i, sku]
    # delta gives chage of sales
    # delta2 gives change in slope (negative means not rising as fast)
    if avg and delta:
        return Avg, Delta
    elif avg:
        return Avg
    elif delta:
        return Delta
### returns average change in sales over last 3 periods, added to last periods sales ###


def AD_forecast(self, period):
    '''produces Forecast using Average Delta method:\n
        given a period, calculates average change in sales over past 3 periods, \nthen adds average change to last periods sales'''
    delta = running_avg(running_avg(self, period, avg=False),
                        period, delta=False).iloc[-1]
    sales = self.iloc[-1:-(period+1):-1].sum()
    return sales.add(delta, fill_value=0).transpose()
### returns sales DF , replacing each week with its total sales over the next period length ###


def DF_grouping(self, period):
    '''groups sales into n-week chunks\n\t'''
    lst = []
    Avg = pd.DataFrame()

    for sku in range(len(self.columns)):
        for i in range(0, self.shape[0]-period+1):
            lst = []
            for j in range(0, period):
                lst.append(self.iloc[i+j, sku])
            Avg.loc[self.index[i], f'{self.columns[sku]}'] = np.sum(lst)
    return Avg
### given 1 year 6 months of sales data, returns each quarters sales as a percentage of the previous quarters sales ###


def seasonality_model(df, period=13):
    '''uses sales from past years to develop a model of sales from period to period based on percentage of sales from 1 period to the next\n
        Returns DF of items relative sales from one period to the next'''
    
    Percentage = pd.DataFrame()

    for sku in range(len(df.columns)):
        for i in range(0, df.shape[0]-2*period):
            lst = []
            lst2 = []
            for j in range(0, period):
                lst2.append(df.iloc[i+j, sku])
            for j in range(period, 2*period):
                lst.append(df.iloc[i+j, sku])
            Percentage.loc[df.index[i+period],
                           f'{df.columns[sku]}'] = np.sum(lst)/np.sum(lst2)
    return Percentage
### uses seasonality_model for projecting 3 months of sales to forecast longer periods, assuming similiar seasonality trends and 2021 ###


def project_Seasonality(self, S_percentage, period):
    '''Deprecated due to inaccuracy\n 
        produces Forecast using Seasonality method:\n\t
        uses relationship between sales from period to period over past year to forecast future sales assuming similiar behavior'''

    s_n = self.grouping(13)
    projection = pd.DataFrame()
    ### for every data point ###
    for sku in range(len(self.columns)):
        for i in range(self.shape[0]-1, self.shape[0]):
            ### returns index in S_pertentage closest to Self's index ###
            time = self.index[i]
            index = S_percentage.index.searchsorted(time)
            try:
                time2 = S_percentage.index[index]
            except:
                time2 = S_percentage.index[0]
            time3 = S_percentage.index[index-1]

            if np.abs(date(2021, int(time2[:2]), int(time2[3:5]))-date(2021, int(time[:2]), int(time[3:5]))) < np.abs(date(2021, int(time3[:2]), int(time3[3:5]))-date(2021, int(time[:2]), int(time[3:5]))):
                a = index
                b = a+13
                c = a+26
            else:
                a = index-1
                b = a+13
                c = a+26
            # checks that index values for S_percentage are within S_percentages's index
            if a >= len(S_percentage.index):
                a = a-len(S_percentage.index)
            if b >= len(S_percentage.index):
                b = b-len(S_percentage.index)
            if c >= len(S_percentage.index):
                c = c-len(S_percentage.index)
            ### multplies last 3 months sales by seasonality muliplier to return next 3 months sales ###
            projection.loc[self.index[i], f'{self.columns[sku]}'] = S_percentage.loc[S_percentage.index[a],
                                                                                     f'{S_percentage.columns[sku]}']*s_n.loc[s_n.index[i-13], f'{s_n.columns[sku]}']
            ### adds next 3 month period -> 6 months ###
            if period == 26:
                projection.loc[self.index[i], f'{self.columns[sku]}'] += projection.loc[self.index[i],
                                                                                        f'{self.columns[sku]}'] * S_percentage.loc[S_percentage.index[b], f'{S_percentage.columns[sku]}']
            ### adds next 6 month period -> 9 months ###
            if period == 39:
                projection.loc[self.index[i], f'{self.columns[sku]}'] += (projection.loc[self.index[i], f'{self.columns[sku]}'] * S_percentage.loc[S_percentage.index[b], f'{S_percentage.columns[sku]}'])+(
                    projection.loc[self.index[i], f'{self.columns[sku]}'] * S_percentage.loc[S_percentage.index[b], f'{S_percentage.columns[sku]}']*S_percentage.loc[S_percentage.index[c], f'{S_percentage.columns[sku]}'])
    return projection


pd.DataFrame.project = project_Seasonality
pd.DataFrame.grouping = DF_grouping
pd.DataFrame.cast = AD_forecast
### formats different length forecasts into single DF ###
### adds columns for predicted order qtys from factory based on 6 months sales ###
### writes data to forecasting google sheet, along with sales data over given periods of time ###


def avg_delta_forecast(S, week_cast, two_week_cast, month_cast):
    '''creates a base or Average Delta Forecast'''
    forecast_df1 = pd.DataFrame()
    three_month_cast = S.cast(13).transpose().apply(lambda x: round(x, 0))
    nine_month_cast = S.cast(26).transpose().apply(lambda x: round(x, 0))
    nine_month_cast = S.cast(39).transpose().apply(lambda x: round(x, 0))

    forecast_df1['1 Week'] = week_cast
    forecast_df1['2 Weeks'] = two_week_cast
    forecast_df1['1 Month'] = month_cast
    forecast_df1['2 Months'] = np.zeros(len(forecast_df1))
    forecast_df1['2 Months'] = ((month_cast.add(three_month_cast))/2).astype(int)
    forecast_df1['3 Months'] = three_month_cast
    forecast_df1['4 Months'] = (nine_month_cast*(2/3)).astype(int)
    forecast_df1['5 Months'] = (nine_month_cast*(5/6)).astype(int)
    forecast_df1['6 Months'] = nine_month_cast
    forecast_df1['7 Months'] = (nine_month_cast*(7/9)).astype(int)
    forecast_df1['8 Months'] = (nine_month_cast*(8/9)).astype(int)
    forecast_df1['9 Months'] = nine_month_cast
    for period in forecast_df1.columns:
        forecast_df1[period][forecast_df1[period] < 0] = 0

    return forecast_df1


def seasonality_forecast(S, S_model, week_cast, two_week_cast, month_cast):
    '''creates a base for Seasonality Forecast'''
    forecast_df2 = pd.DataFrame()
    three_month_project = S.project(S_model, 13).transpose().apply(lambda x: round(x, 0))
    six_month_project = S.project(S_model, 26).transpose().apply(lambda x: round(x, 0))
    nine_month_project = S.project(S_model, 39).transpose().apply(lambda x: round(x, 0))

    forecast_df2['1 Week'] = week_cast
    forecast_df2['2 Weeks'] = two_week_cast
    forecast_df2['1 Month'] = (three_month_project/3).fillna(0).astype(int)
    forecast_df2['2 Months'] = np.zeros(len(forecast_df2))
    forecast_df2['2 Months'] = (((month_cast.add(
        three_month_project.transpose())).transpose().fillna(0))/2).astype(int)
    forecast_df2['3 Months'] = three_month_project
    forecast_df2['4 Months'] = (six_month_project*(2/3)).fillna(0).astype(int)
    forecast_df2['5 Months'] = (six_month_project*(5/6)).fillna(0).astype(int)
    forecast_df2['6 Months'] = six_month_project
    forecast_df2['7 Months'] = (nine_month_project*(7/9)).fillna(0).astype(int)
    forecast_df2['8 Months'] = (nine_month_project*(8/9)).fillna(0).astype(int)
    forecast_df2['9 Months'] = nine_month_project

    return forecast_df2


def ML_forecast(S):
    '''creates a base for Machine Learning Forecast'''
    forecast_df3 = pd.DataFrame(
        np.zeros((77, 11)), index=reduced_sku_list, columns=period_to_weeks.keys())
    for sku in reduced_sku_list:
        for period in period_to_weeks.keys():
            try:
                S.skus_Sales = pd.DataFrame(S[f'{sku}'].values, columns=[
                                            'Sales'], index=S.index)
                S.skus_Sales['Date'] = S.skus_Sales.index
                for index in S.skus_Sales.index:
                    S.skus_Sales.loc[index, 'Month'] = datetime.strptime(
                        S.skus_Sales.loc[index, 'Date'], '%m/%d/%Y').strftime('%m')
                    S.skus_Sales.loc[index, 'Year'] = datetime.strptime(
                        S.skus_Sales.loc[index, 'Date'], '%m/%d/%Y').strftime('%y')
                S.skus_Sales = S.skus_Sales.drop(['Date'], 1)
                S.skus_Sales['Rolling Average'] = S.skus_Sales['Sales'].rolling(
                    3).mean()
                S.skus_Sales['Rolling delta'] = S.skus_Sales['Sales'].rolling(
                    2).apply(lambda x: x.iloc[1] - x.iloc[0])
                S.skus_Sales['avg delta'] = S.skus_Sales['Rolling delta'].rolling(
                    3).mean()
                S.skus_Sales['Rolling delta 2'] = S.skus_Sales['Rolling delta'].rolling(
                    2).apply(lambda x: x.iloc[1] - x.iloc[0])
                S.skus_Sales['avg delta 2'] = S.skus_Sales['Rolling delta 2'].rolling(
                    3).mean()
                S.skus_Sales.fillna(0, inplace=True)
                S.skus_Sales = S.skus_Sales.astype(int)

                pickle_in = open(f'pickles\LR {sku}-{period}.pickle', 'rb')
                clf = pickle.load(pickle_in)

                forecast_df3.loc[sku, period] = (clf.predict(
                    np.array(S.skus_Sales.tail(period_to_weeks[period])))).sum()
            except:
                print(sku, ': ', period)

    return forecast_df3


def Process_Forecast(Sales, queue):
    '''creates 3 forecasts and groupings of sales for several different sized periods'''
    # format and print onto excel sheet
    ### 2 forecast for seperate methods of  forecasting longer periods (3m,6m,9m) ###

    forecast_list1 = []
    forecast_list2 = []
    forecast_list3 = []
    S = Sales.copy()
    S_model = pd.read_json(path+r'\seasonality.json')

    ### totals sales, fill OOS dates, totals to weeks, then takes year off of dates in index ###
    week_cast = S.cast(1).transpose().apply(lambda x: round(x, 0))
    two_week_cast = S.cast(2).transpose().apply(lambda x: round(x, 0))
    month_cast = S.cast(4).transpose().apply(lambda x: round(x, 0))

    forecast_list1 = avg_delta_forecast(
        S, week_cast, two_week_cast, month_cast)
    forecast_list2 = seasonality_forecast(
        S, S_model, week_cast, two_week_cast, month_cast)
    forecast_list3 = ML_forecast(S)

  # FIX WITH QUEUE, add returns to queue, pull from queue here
    # threads = [
    # Thread(target=avg_delta_forecast(S, week_cast, two_week_cast, month_cast,inventory,order_period_dict,weeks_difference,orderdate,orderdate2,priorities)),
    # Thread(target=seasonality_forecast(S,S_model, week_cast, two_week_cast, month_cast,inventory,order_period_dict,weeks_difference,orderdate,orderdate2,priorities)),
    # Thread(target=ML_forecast(S=S))
    # ]
    # n=1
    # for t in threads:
    #     t.start()
    # for t in threads:
    #     exec(f"forecast_list{n} = t.join()")
    #     n+=1

    ### create different forecasts ###
    s2 = S.grouping(2).transpose().apply(lambda x: round(x, 0))
    s4 = S.grouping(4).transpose().apply(lambda x: round(x, 0))
    s13 = S.grouping(13).transpose().apply(lambda x: round(x, 0))
    s26 = S.grouping(26).transpose().apply(lambda x: round(x, 0))
    s39 = S.grouping(39).transpose().apply(lambda x: round(x, 0))

    group1 = [S[-5:].transpose()[-5:].columns.tolist()] + \
        S[-5:].transpose().reset_index().values.tolist()
    group2 = [s2.iloc[:, -5:].columns.tolist()] + s2.iloc[:, -
                                                          5:].reset_index().values.tolist()
    group4 = [s4.iloc[:, -5:].columns.tolist()] + s4.iloc[:, -
                                                          5:].reset_index().values.tolist()
    group13 = [s13.iloc[:, -5:].columns.tolist()] + s13.iloc[:, -
                                                             5:].reset_index().values.tolist()
    group1[0].insert(0, '')
    group2[0].insert(0, '')
    group4[0].insert(0, '')
    group13[0].insert(0, '')
    group26 = [s26.columns.tolist()] + s26.reset_index().values.tolist()
    group39 = [s39.columns.tolist()] + s39.reset_index().values.tolist()
    group26[0].insert(0, '')
    group39[0].insert(0, '')
    group_total = [S.transpose().columns.tolist()] + \
        S.transpose().reset_index().values.tolist()
    group_total[0].insert(0, '')
    global forecasts
    forecasts = [forecast_list1, forecast_list2, forecast_list3,
                 group1, group2, group4, group13, group26, group39, group_total]
    queue.put(forecasts)


def Export_Forecast(exports):
    '''given forecast/grouping list (exports):\n
        uploads to respective Google Forecast Sheets'''
    forecast_list1 = exports[0]
    forecast_list2 = exports[1]
    forecast_list3 = exports[2]
    group1 = exports[3]
    group2 = exports[4]
    group4 = exports[5]
    group13 = exports[6]
    group26 = exports[7]
    group39 = exports[8]
    group_total = exports[9]
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'Avg Delta Forecast!a1',
                          valueInputOption='USER_ENTERED', body={'values': forecast_list1}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'Seasonality Forecast!a1',
                          valueInputOption='USER_ENTERED', body={'values': forecast_list2}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'ML Forecast!a1',
                          valueInputOption='USER_ENTERED', body={'values': forecast_list3}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'1 W history!a1',
                          valueInputOption='USER_ENTERED', body={'values': group1}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'2 W history!a1',
                          valueInputOption='USER_ENTERED', body={'values': group2}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'1 M history!a1',
                          valueInputOption='USER_ENTERED', body={'values': group4}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'3 M history!a1',
                          valueInputOption='USER_ENTERED', body={'values': group13}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'6 M history!a1',
                          valueInputOption='USER_ENTERED', body={'values': group26}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'9 M history!a1',
                          valueInputOption='USER_ENTERED', body={'values': group39}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'sales history total!a1',
                          valueInputOption='USER_ENTERED', body={'values': group_total}).execute()


def Forecast_graphing(sales, forecasts, period, sku):
    '''DEPRECATED\n
        creates a graph with real sales followed by forecasted sales (following a lilnear trend from last weeks sales)'''
    i = 1
    for forecast in forecasts:
        P = period_to_weeks[period]
        N = forecast.loc[sku, period]
        n1 = sales[sku][-1:][0]
        n1_date = sales.index[-1].split('/')
        n1_date = date(int(n1_date[2]), int(n1_date[0]), int(n1_date[1]))
        dates_list = []
        values_list = []
        for T in range(P+1):
            dates_list.append(
                (n1_date+relativedelta(weeks=T)).strftime("%m/%d/%Y"))
            values_list.append(n1+((2*T/P)*((N/P)-n1)))
        if i == 1:
            df = pd.DataFrame(values_list, index=dates_list)
            sales.sku = sales[sku]
            for j in range(len(dates_list)):
                sales.sku = sales.sku.append(
                    pd.Series(np.nan, index=[dates_list[j]]))
        else:
            df = pd.merge(df, pd.DataFrame(values_list, index=dates_list),
                          how='inner', left_index=True, right_index=True)
        i += 1

        # sales.index[-1]=dates_list[j]
    sales.sku = pd.DataFrame(
        sales.sku.values, index=sales.sku.index.values.tolist(), columns=[sku])
    sales.sku = pd.merge(sales.sku, df, how='left',
                         left_index=True, right_index=True)
    sales.sku.columns = ["sales", "Avg Delta", "Seasonality", "ML"]
    return sales.sku


def process_sales(queue):
    '''uses exported sales data and applies 3 proesses to return DF with sales grouped by week, and with OOS periods removed'''
    global processed_sales
    s = sales(path+rf'\exports\Sales Data.xlsx')
    unprocessed_sales = weekify(s)
    processed_sales = weekify(remove_OOS(s))
    queue.put((unprocessed_sales, processed_sales))


def Production_Schedule_0(page):
    '''DEPRECATED\n
        builds production schedules and uploads to google sheets'''
    ### plug in settings ###
    settings = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                            range=f"'production schedule settings'!a1:b100").execute().get('values', []))
    max_containers = int(settings.loc[0, 1])
    mipc = int(settings.loc[1, 1])

    def myround(x, base=(1/mipc)):
        return base * round(x/base)
    skus = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                        range=f"'production schedule {page}'!b2:j80").execute().get('values', []))
    skus = skus.drop(0)
    ### grab values from GS and add to DF ###
    df1 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                       range=f"'production schedule {page}'!k2:q80").execute().get('values', []))
    df1 = df1.drop(0)
    df1.columns = ['priority', 'qty', 'containers', 'AD load plan 1',
                   'AD load plan 2', 'AD load plan 3', 'AD load plan 4']
    df1[['sku', 'max order']] = skus[[0, 6]].values
    df1['priority'] = df1['priority'].astype(int)
    df1 = df1.sort_values(by='priority', ascending=False)

    df2 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                       range=f"'production schedule {page}'!r2:x80").execute().get('values', []))
    df2 = df2.drop(0)
    df2.columns = ['priority', 'qty', 'containers', 'AD load plan 1',
                   'AD load plan 2', 'AD load plan 3', 'AD load plan 4']
    df2[['sku', 'max order']] = skus[[0, 7]].values
    df2['priority'] = df2['priority'].astype(int)
    df2 = df2.sort_values(by='priority', ascending=False)

    df3 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                       range=f"'production schedule {page}'!y2:ae80").execute().get('values', []))
    df3 = df3.drop(0)
    df3.columns = ['priority', 'qty', 'containers', 'AD load plan 1',
                   'AD load plan 2', 'AD load plan 3', 'AD load plan 4']
    df3[['sku', 'max order']] = skus[[0, 8]].values
    df3['priority'] = df3['priority'].astype(int)
    df3 = df3.sort_values(by='priority', ascending=False)

    ### make containers int instead of str ###
    df1['containers'] = df1['containers'].astype(float)
    df2['containers'] = df2['containers'].astype(float)
    df3['containers'] = df3['containers'].astype(float)

    ### creates loading plans by adding 1/4 of most needed items first, until were out of room  ###
    ### Avg Delta ###
    for method in [['AD', df1, 'n2'], ['S', df2, 'u2'], ['Mid', df3, 'ab2']]:
        containers = method[1]['containers'].astype(float).tolist()
        orders = np.zeros((4, method[1].shape[0])).tolist()
        max_order = method[1]['max order'].tolist()
        for i in range(len(containers)):
            if max_order[i] in [None, ""]:
                max_order[i] = containers[i]
            elif max_order[i].upper() == 'X':
                containers[i] = 0
            elif containers[i] > float(max_order[i]):
                containers[i] = float(max_order[i])
        method[1]['containers'] = containers
        for i in range(4):
            for j in range(len(orders[i])):
                skus_containers = 0
                for k in range(4):
                    skus_containers += orders[k][j]
                if sum(orders[i]) < max_containers and (containers[j]/4) <= 0.25 and skus_containers+myround(containers[j]/2) <= containers[j]:
                    orders[i][j] = myround(containers[j])
                elif sum(orders[i]) < max_containers and skus_containers+myround(containers[j]/4) <= containers[j]:
                    orders[i][j] = myround(containers[j]/4)
                elif sum(orders[i]) < max_containers:
                    orders[i][j] = containers[j]-skus_containers

        ### assign order values to df ###
        method[1][f'{method[0]} load plan 1'] = orders[0]
        method[1][f'{method[0]} load plan 2'] = orders[1]
        method[1][f'{method[0]} load plan 3'] = orders[2]
        method[1][f'{method[0]} load plan 4'] = orders[3]
        ### set type as float, reorder index, add/subtract leftover containers, round to nearest 0.25 ###
        method[1][[f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']] = method[1][[
            f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']].astype(float)
        method[1] = method[1].sort_index()
        method[1]['containers left'] = method[1]['containers'] - \
            method[1][f'{method[0]} load plan 1']-method[1][f'{method[0]} load plan 2'] - \
            method[1][f'{method[0]} load plan 3'] - \
            method[1][f'{method[0]} load plan 4']
        method[1][f'{method[0]} load plan 4'] += method[1]['containers left']
        # method[1]['containers left']=method[1]['containers']-method[1][f'{method[0]} load plan 1']-method[1][f'{method[0]} load plan 2']-method[1][f'{method[0]} load plan 3']-method[1][f'{method[0]} load plan 4']
        method[1][[f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']] = method[1][[
            f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']].apply(myround)
        method[1][f'{method[0]} load plan 3'] += np.where(
            method[1][f'{method[0]} load plan 4'] < 0, method[1][f'{method[0]} load plan 4'], 0)
        method[1][f'{method[0]} load plan 4'] = np.where(
            method[1][f'{method[0]} load plan 4'] < 0, 0, method[1][f'{method[0]} load plan 4'])
        method[1][f'{method[0]} load plan 4'] = np.where(
            method[1][f'{method[0]} load plan 3'] == 0, 0, method[1][f'{method[0]} load plan 4'])

        method[1] = method[1][[f'{method[0]} load plan 1', f'{method[0]} load plan 2',
                               f'{method[0]} load plan 3', f'{method[0]} load plan 4']]
        df_list = [method[1].columns.tolist()] + method[1].values.tolist()
        sheet.values().update(spreadsheetId=FORECAST_ID,
                              range=f"'production schedule {page}'!{method[2]}", valueInputOption='USER_ENTERED', body={'values': df_list}).execute()


def read_forecasts():
    '''reads forecast data from forecast google sheets\n
        Returns as list of DataFrames'''
    df1 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                       range="'Avg Delta Forecast'!A2:L100").execute().get('values', []))
    df2 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                       range="'Seasonality Forecast'!A2:L100").execute().get('values', []))
    df3 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                       range="'ML Forecast'!A2:L100").execute().get('values', []))
    for df in [df1, df2, df3]:
        df.index = df[0]
        df.drop(0, axis=1, inplace=True)
        # df.columns=df.iloc[0]
        df.columns = ['1 Week', '2 Weeks', '1 Month', '2 Months', '3 Months',
                      '4 Months', '5 Months', '6 Months', '7 Months', '8 Months', '9 Months']
        # df=df[['1 Week','2 Weeks','1 Month','2 Months','3 Months','4 Months','5 Months','6 Months','7 Months','8 Months','9 Months']].astype(float)
    return[df1.astype(float), df2.astype(float), df3.astype(float)]


def read_sales():
    '''reads sales data from forecast google sheets\n
        Returns as DataFrame'''
    df = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID,
                      range="'sales history total'!A1:zz100").execute().get('values', []))
    columns = df[0][1:]
    df = df.drop(0, axis=1).transpose()
    df.index = df[0]
    df = df.drop(0, axis=1)
    df.columns = columns
    return df.astype(float)


def Days_in_stock(*args):
    '''creates DF of skus velocities, simulated velocities, and divides current stock by them to give days in stock\n
        uploads to google sheet'''
    df = pd.read_excel(path+r'\exports\Velocities.xlsx')
    df.index = df['ID']
    df = df[["Velocity"]]
    df
    velocities = {}
    for sku in reduced_sku_list:
        try:
            velocities[sku] = df.loc[sku, "Velocity"]
        except:
            velocities[sku] = 0
    df2 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID, range="'1 M history'!A2:f100").execute(
    ).get('values', []), columns=['sku', 'x1', 'x2', 'x3', 'x4', 'current'])
    df3 = pd.DataFrame(sheet.values().get(spreadsheetId=FORECAST_ID, range="'Avg Delta Forecast'!A2:O100").execute().get(
        'values', []), columns=['sku', 'x1', 'x2', 'x3', 'x4', 'x5', 'x6', 'x7', 'x8', 'x9', 'x10', 'x11', 'x12', 'x13', 'Stock'])
    # df3 = item_quantity()
    df2['Stock'] = df3[['Stock']]
    df = pd.DataFrame(velocities.values(),
                      index=velocities.keys(), columns=["Velocity"])
    df['Sim. Velocity'] = 0
    df['Stock'] = 0
    for i in df2.itertuples(index=False):
        try:
            df.loc[i[0], 'Sim. Velocity'] += int(i[5])/30
            if int(i[6]) > 0:
                df.loc[i[0], 'Stock'] = int(i[6])
        except:
            pass
    df['Days left in stock'] = df['Stock']/df['Velocity']
    df['Days left in stock (simulated)'] = df['Stock']/df['Sim. Velocity']
    df = df.fillna(0).apply(lambda x: round(x, 1))
    df.replace([np.inf, -np.inf], '', inplace=True)
    df = [df.columns.tolist()] + df.reset_index().values.tolist()
    df[0].insert(0, 'SKU')
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'days in stock!a1',
                          valueInputOption='USER_ENTERED', body={'values': df}).execute()
    sheet.values().update(spreadsheetId=FORECAST_ID, range=f'days in stock!h1',
                          valueInputOption='USER_ENTERED', body={'values': [[datetime.today().strftime("%D")]]}).execute()


def download_files(queue):
    '''using threads, downloads all needed Seller Cloud Exports\n
     waits until all processes are finished'''
    chrome_path=ChromeDriverManager().install()
    threads = [Thread(target=download_sales, args=(chrome_path,)),Thread(target=download_vel, args=(chrome_path,)),Thread(target=download_inv, args=(chrome_path,))]
    for t in threads:
        t.start()
    for t in threads:
        t.join()


def prepare_exports(Sales, forecasts):
    '''given bases of forecasts,\n
        adds total inventory column\n
        adds order quantities for every month up to 9 months out\n
        returns Forecast DFs '''
    inventory = item_quantity()
    order_period_dict = {'1 Month': 2, '2 Months': 3, '3 Months': 4, '4 Months': 5,
                         '5 Months': 6, '6 Months': 7, '7 Months': 8, '8 Months': 9, '9 Months': 10}
    S = Sales.copy()
    ### set up to calculate dales extra sales due to time between now and order date ###
    priorities = pd.read_json(path+r'\priority.json')
    day = date(int(S.index[-1][6:]), int(S.index[-1][:2]),
               int(S.index[-1][3:5]))+timedelta(weeks=1)
    orderdate = date(day.year, day.month, 15)
    orderdate2 = date(day.year, day.month+1, 15)
    weeks_difference = (orderdate-day).days/7
    exports = []
    for i in range(3):
        forecast = forecasts[i].copy()
        forecast['Priority'] = 0
        for period in forecast.columns[:-1]:
            forecast[period][forecast[period] < 0] = 0
        forecast[''] = ''
        forecast['Stock'] = inventory

        for order_period in order_period_dict.keys():
            filler_qty = (forecast[forecast.columns[order_period_dict[order_period]+1]
                                   ]-forecast[order_period])*(weeks_difference/4)
            if weeks_difference < 0:
                filler_qty = weeks_difference*forecast['1 Month']/4
            forecast[f'{order_period}: {orderdate.strftime("%m/%d")} Order QTY'] = (
                forecast[order_period]+filler_qty.transpose()-forecast['Stock']).transpose().apply(lambda x: round(x, 0))
            forecast[f'{order_period}: {orderdate2.strftime("%m/%d")} Order QTY'] = (
                (forecast[forecast.columns[order_period_dict[order_period]+1]]-forecast[order_period])).apply(lambda x: round(x, 0))

        forecast['Priority'] = np.where(
            forecast['9 Months'] < forecast['Stock'], 0, forecast['Priority'])
        forecast['Priority'] = np.where(
            forecast['9 Months'] > forecast['Stock'], 1, forecast['Priority'])
        forecast['Priority'] = np.where(
            forecast['6 Months'] > forecast['Stock'], 2, forecast['Priority'])
        forecast['Priority'] = np.where(
            forecast['3 Months'] > forecast['Stock'], 3, forecast['Priority'])
        forecast['Priority'] = np.where(
            forecast['2 Months'] > forecast['Stock'], 4, forecast['Priority'])
        forecast['Priority'] = (forecast['Priority'].astype(
            int)).multiply(priorities["Last 12 mos. Sales"])
        forecast = forecast.fillna(0)

        forecast_list = [forecast.columns.tolist()] + \
            forecast.reset_index().values.tolist()
        forecast_list[0].insert(0, str(S.index[-1:][0]))

        exports.append(forecast_list)
    for i in range(3, len(forecasts)):
        exports.append(forecasts[i])
    return exports


def Export_Production_Schedule(export):
    '''given production schedule from Prepare_Production_Schedule(), exports and formats as an excel file'''
    arrays = [export['factory code'].tolist(), export.index.tolist()]
    tuples = list(zip(*arrays))
    index = pd.MultiIndex.from_tuples(tuples, names=["Factory Code", "SKU"])
    attempt1 = pd.DataFrame(export.values, index=index,
                            columns=export.columns.tolist())
    attempt1.drop(axis=1, labels='factory code', inplace=True)
    attempt1.sort_index(axis=0, inplace=True)
    attempt1.to_excel(
        f'production export/production schedule {datetime.today().strftime("%m-%d")}.xlsx')
    style_export()


def Prepare_Production_Schedule(sku_details, period, factories, in_production, max_containers):
    '''creates production schedule'''
    #* produces schedule up until 9mo-leadtime
    
    ### list of next 6 months ###
    months_in_year = ['January', 'February', 'March', 
    'April', 'May', 'June', 
    'July', 'August', 'September', 
    'October', 'November', 'December',
    'January', 'February', 'March', 
    'April', 'May', 'June',
    'July', 'August', 'September', 
    'October', 'November', 'December']
    months_general=['cur', 1,2,3,4,5,6,7]
    nine_mo = []
    cur_month = datetime.today().month
    for i in range(10):
        nine_mo.append([months_in_year[cur_month+i]])
    ### plug in settings ###

    if max_containers == '':
        max_containers = 1000
    mipc = 4  # int(settings.loc[1,1]) #4

    def myround(x, base=(1/mipc)):
        return base * round(x/base)
    ### grab values from GS and add to DF ###
    # probably wont work, placeholder
    forecasts = read_forecasts()
    forecast = (forecasts[0]+forecasts[2])/2
    ip = (pd.read_excel(in_production)).set_index('sku')
    # inv = item_quantity()['Stock']
    forecast['Stock'] = item_quantity()['Stock']

    day = datetime.date(datetime.today())
    orderdate = date(day.year, day.month, 15)
    weeks_difference = (orderdate-day).days/7
    filler_qty = (forecast['1 Month'])*(weeks_difference/4)

    sku_details = pd.DataFrame(sku_details).transpose()

    def meta(forecast, filler_qty, method):
        forecast.meta = forecast.copy()
        if nine_mo.index(method)!=7:
            Scheduled_qty=ip[f'{months_general[nine_mo.index(method)]} mo qty']
            forecast['Stock'] = forecast.meta['Stock'].add(Scheduled_qty, fill_value=0)
        
        forecast.meta['qty'] = (forecast[period]+filler_qty)-forecast['Stock'] 
        # forecast.meta['qty'][forecast.meta['qty'] < 0] = 0 #TEMP: commented to allow for negative qty (if we have more than enough inventory)
        priorities = pd.read_json(path+r'\priority.json')
        forecast.meta['Priority'] = 0
        forecast.meta['Priority'] = np.where(
            forecast.meta['9 Months'] < forecast.meta['Stock'], 0, forecast.meta['Priority'])
        forecast.meta['Priority'] = np.where(
            forecast.meta['9 Months'] > forecast.meta['Stock'], 1, forecast.meta['Priority'])
        forecast.meta['Priority'] = np.where(
            forecast.meta['6 Months'] > forecast.meta['Stock'], 2, forecast.meta['Priority'])
        forecast.meta['Priority'] = np.where(
            forecast.meta['3 Months'] > forecast.meta['Stock'], 3, forecast.meta['Priority'])
        forecast.meta['Priority'] = np.where(
            forecast.meta['2 Months'] > forecast.meta['Stock'], 4, forecast.meta['Priority'])
        forecast.meta['Priority'] = (forecast.meta['Priority'].astype(
            int)).multiply(priorities["Last 12 mos. Sales"])
        forecast.meta = forecast.meta.fillna(0)

        forecast.meta['Priority'] = forecast.meta['Priority'].astype(int)
        forecast.meta = forecast.meta.sort_values(
            by='Priority', ascending=False)

        ### make containers int instead of str ###
        forecast.meta['sku'] = forecast.index

        forecast.meta[f'{method[0]} containers'] = (forecast.meta['qty'].astype(
            float).divide(sku_details['container qty'], fill_value=1)).apply(lambda x: myround(x))
        forecast.meta[f'{method[0]} load plan 1'] = 0
        forecast.meta[f'{method[0]} load plan 2'] = 0
        forecast.meta[f'{method[0]} load plan 3'] = 0
        forecast.meta[f'{method[0]} load plan 4'] = 0

        forecast.meta = forecast.meta[['Priority', 'qty', f'{method[0]} containers', f'{method[0]} load plan 1',
                                       f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4', 'sku']]
        forecast.meta['factory code'] = sku_details['factory code']

        def factory_check(item):
            return item in factories

        forecast.meta['fillable'] = forecast.meta['factory code'].map(factory_check)
        forecast.meta[f'{method[0]} containers'][forecast.meta['fillable'] == False] = 0

        return forecast.meta

    export = pd.DataFrame()
    export.index = forecast.index
    export['factory code'] = sku_details['factory code']
    export['Reliability'] = sku_details['forecastability']

    for method in nine_mo[:9-period_num[period]]:
        cur_df = meta(forecast, filler_qty, method)
        containers = cur_df[f'{method[0]} containers'].astype(float).tolist()
        orders = np.zeros((4, cur_df.shape[0])).tolist()
        
        # max_order=cur_df['max order'].tolist()
        # for i in range(len( containers)):
        #     if max_order[i]in [None,""]:
        #         max_order[i]=containers[i]
        #     elif max_order[i].upper() == 'X':
        #         containers[i]=0
        #     elif containers[i]>float(max_order[i]):
        #         containers[i]=float(max_order[i])
        cur_df[f'{method[0]} containers'] = containers
        for i in range(4):
            for j in range(len(orders[i])):
                skus_containers = 0
                for k in range(4):
                    skus_containers += orders[k][j]
                if containers[j]<0:
                    pass
                elif sum([n for n in orders[i] if n>=0]) < max_containers and (containers[j]/4) <= 0.25 and skus_containers+myround(containers[j]/2) <= containers[j]: #(containers[j]/4) <= 0.25 and np.abs(skus_containers+myround(containers[j]/2)) < np.abs(containers[j]):
                    orders[i][j] = myround(containers[j])
                elif sum([n for n in orders[i] if n>=0]) < max_containers and skus_containers+myround(containers[j]/4) <= containers[j]:
                    orders[i][j] = myround(containers[j]/4)
                elif sum([n for n in orders[i] if n>=0]) < max_containers:
                    orders[i][j] = containers[j]-skus_containers

        ### assign order values to df ###
        cur_df[f'{method[0]} load plan 1'] = orders[0]
        cur_df[f'{method[0]} load plan 2'] = orders[1]
        cur_df[f'{method[0]} load plan 3'] = orders[2]
        cur_df[f'{method[0]} load plan 4'] = orders[3]
        ### set type as float, reorder index, add/subtract leftover containers, round to nearest 0.25 ###
        cur_df[[f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']] = cur_df[[
            f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']].astype(float)
        cur_df = cur_df.sort_index()
        cur_df[f'{method[0]} containers left'] = cur_df[f'{method[0]} containers']-cur_df[f'{method[0]} load plan 1'] - \
                                                                                    cur_df[f'{method[0]} load plan 2'] - \
                                                                                    cur_df[f'{method[0]} load plan 3'] - \
                                                                                    cur_df[f'{method[0]} load plan 4']
        
        
        #FIXME fucing up negative container qtys, then doubling plan2 
        cur_df[f'{method[0]} load plan 4'] += cur_df[f'{method[0]} containers left']
        # cur_df[f'{method[0]} containers left']=cur_df[f'{method[0]} containers']-cur_df[f'{method[0]} load plan 1']-cur_df[f'{method[0]} load plan 2']-cur_df[f'{method[0]} load plan 3']-cur_df[f'{method[0]} load plan 4']
        
        cur_df[[f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']] = cur_df[[
            f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3', f'{method[0]} load plan 4']].apply(myround)
        # if plan4<0 and others arent, subtract from plan3
        cur_df[f'{method[0]} load plan 3'] += np.where((
            cur_df[f'{method[0]} load plan 4'] < 0) & ((cur_df[f'{method[0]} load plan 1'] +cur_df[f'{method[0]} load plan 2'] +cur_df[f'{method[0]} load plan 3'] )> 0),
            cur_df[f'{method[0]} load plan 4'], 0)
        # if plan3<0 and others arent, subtract from plan2
        cur_df[f'{method[0]} load plan 2'] += np.where(
            cur_df[f'{method[0]} load plan 3'] < 0, cur_df[f'{method[0]} load plan 3'], 0)
        # if plan4<0 and others arent, replace w/ 0
        cur_df[f'{method[0]} load plan 4'] = np.where((
            cur_df[f'{method[0]} load plan 4'] < 0) & ((cur_df[f'{method[0]} load plan 1'] +cur_df[f'{method[0]} load plan 2'] +cur_df[f'{method[0]} load plan 3'] )> 0),
             0, cur_df[f'{method[0]} load plan 4'])
        # if plan3<0 and others arent, replace w/ 0
        cur_df[f'{method[0]} load plan 3'] = np.where((
            cur_df[f'{method[0]} load plan 3'] < 0) & ((cur_df[f'{method[0]} load plan 1'] +cur_df[f'{method[0]} load plan 2'])> 0),
             0, cur_df[f'{method[0]} load plan 3'])
        # if plan3 == 0, replace plan 4 with 0 as well
        cur_df[f'{method[0]} load plan 4'] = np.where(
            cur_df[f'{method[0]} load plan 3'] == 0, 0, cur_df[f'{method[0]} load plan 4'])

        # sum load plans for total containers column
        cur_df[f'{method[0]} containers'][cur_df[f'{method[0]} containers']>0] = cur_df[f'{method[0]} load plan 1'] + cur_df[f'{method[0]} load plan 2'] + cur_df[f'{method[0]} load plan 3'] + cur_df[f'{method[0]} load plan 4']
        # add ordered qtys to "current stock", assuming schedule followed strictly
        cur_df['qty'] = cur_df[f'{method[0]} containers'][cur_df[f'{method[0]} containers']>0].multiply(sku_details['container qty'])
        forecast['Stock'] = forecast['Stock'].add(cur_df['qty'], fill_value=0)
        cur_df = cur_df[[f'{method[0]} load plan 1', f'{method[0]} load plan 2',
                         f'{method[0]} load plan 3', f'{method[0]} load plan 4', f'{method[0]} containers']]
        # add columns to export df
        export[[f'{method[0]} load plan 1', f'{method[0]} load plan 2', f'{method[0]} load plan 3',
                f'{method[0]} load plan 4', f'{method[0]} containers']] = cur_df
        # change to next months schedule
        period = num_period[period_num[period]+1]

    return export


def style_export():
    '''styles production schedule'''
    wkbk = xl.load_workbook(
        f'production export/production schedule {datetime.today().strftime("%m-%d")}.xlsx')
    obj = wkbk.active
    col_len = obj.max_row+1
    yellow = xl.styles.colors.Color(rgb='FFFF00')
    black = xl.styles.colors.Color(indexed=0)
    bluefill = xl.styles.PatternFill(fill_type='solid', start_color=yellow)
    blackfill = xl.styles.PatternFill(fill_type='solid', start_color=black)
    row = 0
    set_val = 'Factory Code'
    ### unmerge A columns ###
    for cell_group in obj.merged_cell_ranges:
        obj.unmerge_cells(str(cell_group))
    ### fill in Factory codes in column A ###
    while True:
        row += 1
        val = obj.cell(row, 1).value
        if (val != None) & (row != 1):
            obj.insert_rows(row, 1)
            row += 1
            col_len += 1
            set_val = val
        else:
            if (obj.cell(row, 2).value == None) & (set_val == 'disregard'):
                break
            else:
                obj.cell(row, 1).value = set_val
    counter = 0
    ### fill total container columns blue ###
    for col in obj.columns:
        counter += 1
        if ((counter-8) % 5 == 0) & (counter != 2):
            for cell in col:
                cell.fill = bluefill
    # fill rows between factories black
    for row in obj.rows:
        if (row[0].value == None):
            for cell in row:
                cell.fill = blackfill
    ### style as table ###
    if obj.max_column<27:
        maxcol = alphas[obj.max_column-1]
    else:
        maxcol = "A"+alphas[obj.max_column-27]
    tab = Table(displayName="Table1", ref=f"A1:{maxcol}90")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    tab.tableStyleInfo = style
    
    obj.add_table(tab)
    for _row in obj.rows:
        for _cell in _row:
            _cell.border = thin_border
    column_widths = []
    for row in obj.columns:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = len(str(cell))
            else:
                column_widths += [len(str(cell))]

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        obj.column_dimensions[get_column_letter(i)].width = column_width

    wkbk.save(
        f'production export/production schedule {datetime.today().strftime("%m-%d")}.xlsx')
